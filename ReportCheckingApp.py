#===============================================================================
# Report Checking Application
#===============================================================================
'''
Based on      =>    Easy U_2.5 EMC Test Plan For Customer
Project       =>    ECMT2 A5 Validation Testing and Diagnosing

SVN File      =>    https://192.168.5.65/svn/PECMTPE1/Development/Hardware/
Validation/TestPlan/EasyU_2.5 EMC test plan for customer.pdf

Written By    =>    Omid Seyedian
Maintainer    =>    Omid Seyedian
Group         =>    PTHW
Date&Time     =>    Tuesday, June 8, 2021
Location      =>    CROUSE, Saloon 2B, 1st Floor, R&I Department
Written in    =>    Python 3.9.4
'''

#===============================================================================
# Libraries And Modules
#===============================================================================
from _operator import concat, abs                   # @UnusedImport
from builtins import dict
import datetime
from idlelib.idle_test.test_editor import insert    # @UnusedImport
from math import floor
import math
import os
import string
import threading
import time

from openpyxl import load_workbook
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, Border
from openpyxl.styles import Side, colors
from statistics import mean
import statistics
import svn.local

#---------------------------------------------------------------- TicToc Prince!
very_start_time = time.time()

#===============================================================================
# Constant Parameters
#===============================================================================
# Number of Analog Input Parameters
ANALOG_INPUTS = 25

# Digital inputs
DIGITAL_INPUTS = 8

# Frequency inputs
FREQ_INPUTS = 10
TLE_REGISTERS = 8
ETC_REGISTERS = 8
KP254_VALUES = 4
TEMPERATURE_VALUES = 4

# Reset registers
RESET_REGISTERS = 5

CHART_HIGHT = 15
CHART_WIDTH = 24.2705

#------------------------------------------------------------ Standard Tolerance
TOLERANCE = 0.05

#------------------------------------------------------------------------ Colors
VIOLET = openpyxl.styles.colors.Color(rgb='D76EEC')
VIOLET_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=VIOLET)

PEACH = openpyxl.styles.colors.Color(rgb='CCFFFF')
PEACH_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=PEACH)

BLUE = openpyxl.styles.colors.Color(rgb='FFCC99')
BLUE_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=BLUE)

LGREY = openpyxl.styles.colors.Color(rgb='F8F8F8')
LGREY_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=LGREY)

GREY = openpyxl.styles.colors.Color(rgb='C0C0C0')
GREY_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=GREY)

RED = openpyxl.styles.colors.Color(rgb='FF0000')
RED_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=RED)

GREEN = openpyxl.styles.colors.Color(rgb='99FF99')
GREEN_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=GREEN)

YELLOW = openpyxl.styles.colors.Color(rgb='FFFF66')
YELLOW_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=YELLOW)

BROWN = openpyxl.styles.colors.Color(rgb='CC9900')
BROWN_FILL = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=BROWN)

#--------------------------------------------------- Excel Standard Column Names
litterals = list(string.ascii_uppercase)
Excel_std_column_names = litterals.copy()
litterals_aid = litterals.copy()

for x in litterals:
    for y in litterals_aid:
        Excel_std_column_names.append(x + y)

#===============================================================================
# Functions
#===============================================================================



def parameters_function():
    '''
    Creating parameters dictionary.

    This function creates a dictionary between
    the main sheet's columns and their parameter.
    '''
    parameters = []
    for x in main_sheet.iter_rows(min_row=8,
                                  max_row=8,
                                  min_col=2,
                                  max_col=EXCEL_HRZN_SCOPE,
                                  values_only=True):
        parameters.append(x)
    numerations = tuple([x for x in range(EXCEL_HRZN_SCOPE)])
    pure_dictionary = dict(zip(parameters[0], numerations))
    pure_parameters = tuple(pure_dictionary.keys())

    columns = [x for x in range(1, 131) if x % 3 == 0]
    columns += [x for x in range(130, 144)]
    columns += [x for x in range(144, 158)]
    columns += [x for x in range(158, 159)]
    columns += [x for x in range(174, 175)]
    columns = tuple(columns)

    parameters_dict = dict(zip(pure_parameters, columns))
    return parameters_dict



def analog_service_pack(parameter):
    '''
    This function processes all of the analog parameters.
    '''
    parameter_column = parameters_function()[parameter]
    service_result = [[], []]
    values = []
    for x in main_sheet.iter_rows(min_row=10,
                                  max_row=EXCEL_VERT_SCOPE,
                                  min_col=parameter_column,
                                  max_col=parameter_column + 1,
                                  values_only=True):
        values.append(x[0])
        values.append(x[1])

    minimum_value = round(min(values), 3)
    maximum_value = round(max(values), 3)
    average_value = round(mean(values), 3)
    standard_deviation = round(math.sqrt(statistics.variance(values)), 3)

    service_result[0].append('Parameter')
    service_result[0].append('Minimum Value [V]')
    service_result[0].append('Maximum Value [V]')
    service_result[0].append('Average Value [V]')
    service_result[0].append('Standard Deviation [V]')

    service_result[1].append(parameter)
    service_result[1].append(minimum_value)
    service_result[1].append(maximum_value)
    service_result[1].append(average_value)
    service_result[1].append(standard_deviation)

    if average_value != 0:
        coef_variation = round(100 * standard_deviation / average_value, 3)
        service_result[0].append('Coef. Variation [%]')
        service_result[1].append(coef_variation)

    if (parameter in constant_issued_values.keys()
        and
            constant_issued_values[parameter] != '??'):
        avg_error = round(100 * (constant_issued_values[parameter] \
                                 -average_value) \
                                 / constant_issued_values[parameter], 2)
        service_result[0].append('Average Error [%]')
        service_result[1].append(avg_error)
        service_result[0].insert(1, 'Issued Value')
        service_result[1].insert(1, constant_issued_values[parameter])
    return service_result



def byte_to_pressure(doublet):
    '''
    This function computes the pressure value, from digital bytes.

    Argument: A 16bit digital value
    return  : Pressure
    '''
    msb = list(doublet[:8])
    lsb = list(doublet[8:])

    while len(msb) < 8:
        msb.insert(0, '0')

    while len(lsb) < 8:
        lsb.insert(0, '0')

    decimal_value = 1 * int(lsb[-2]) + \
                    2 * int(lsb[-3]) + \
                    4 * int(lsb[-4]) + \
                    8 * int(lsb[-5]) + \
                    16 * int(lsb[-6]) + \
                    32 * int(lsb[-7]) + \
                    64 * int(lsb[-8]) + \
                    128 * int(msb[-1]) + \
                    256 * int(msb[-2]) + \
                    512 * int(msb[-3])

    pressure = round((decimal_value + 545.6) / 13.64, 2)
    return pressure



def byte_to_temperature(doublet):
    '''
    This function computes the temperature value, from digital bytes.

    Argument: A 16bit digital value
    return  : Temperature
    Dear one, please refer to the 11'th page of the data sheet.
    '''
    msb = list(doublet[:8])
    lsb = list(doublet[8:])

    while len(msb) < 8:
        msb.insert(0, '0')

    while len(lsb) < 8:
        lsb.insert(0, '0')

    decimal_value = 1 * int(lsb[-2]) + \
                    2 * int(lsb[-3]) + \
                    4 * int(lsb[-4]) + \
                    8 * int(lsb[-5]) + \
                    16 * int(lsb[-6]) + \
                    32 * int(lsb[-7]) + \
                    64 * int(lsb[-8]) + \
                    128 * int(msb[-1]) + \
                    256 * int(msb[-2]) + \
                    512 * int(msb[-3])

    temperature = round((decimal_value - 204.6) / 5.115)
    return temperature



def KP254_diagnosis(sheet, _row, _column, high_diag_btye):
    diag = high_diag_btye[:5]

    if diag == '10000':
        sheet.cell(row=_row,
                   column=_column).value = 'FEC error'
        sheet.cell(row=_row, column=_column).fill = RED_FILL
        sheet.cell(row=1, column=_column).fill = RED_FILL
        main_sheet.cell(row=7, column=144).fill = RED_FILL
        main_sheet.cell(row=7, column=145).fill = RED_FILL
        main_sheet.cell(row=7, column=146).fill = RED_FILL
        main_sheet.cell(row=7, column=147).fill = RED_FILL

    elif diag == '01000':
        sheet.cell(row=_row,
                   column=_column).value = 'Acquisition chain failure'
        sheet.cell(row=_row, column=_column).fill = RED_FILL
        sheet.cell(row=1, column=_column).fill = RED_FILL
        main_sheet.cell(row=7, column=144).fill = RED_FILL
        main_sheet.cell(row=7, column=145).fill = RED_FILL
        main_sheet.cell(row=7, column=146).fill = RED_FILL
        main_sheet.cell(row=7, column=147).fill = RED_FILL

    elif diag == '00100':
        sheet.cell(row=_row,
                   column=_column).value = 'Sensor cell failure'
        sheet.cell(row=_row, column=_column).fill = RED_FILL
        sheet.cell(row=1, column=_column).fill = RED_FILL
        main_sheet.cell(row=7, column=144).fill = RED_FILL
        main_sheet.cell(row=7, column=145).fill = RED_FILL
        main_sheet.cell(row=7, column=146).fill = RED_FILL
        main_sheet.cell(row=7, column=147).fill = RED_FILL

    elif diag == '00010':
        sheet.cell(row=_row,
                   column=_column).value = 'Pressure out of range (High)'
        sheet.cell(row=_row, column=_column).fill = RED_FILL
        sheet.cell(row=1, column=_column).fill = RED_FILL
        main_sheet.cell(row=7, column=144).fill = RED_FILL
        main_sheet.cell(row=7, column=145).fill = RED_FILL
        main_sheet.cell(row=7, column=146).fill = RED_FILL
        main_sheet.cell(row=7, column=147).fill = RED_FILL

    elif diag == '00001':
        sheet.cell(row=_row,
                   column=_column).value = 'Pressure out of range (Low)'
        sheet.cell(row=_row, column=_column).fill = RED_FILL
        sheet.cell(row=1, column=_column).fill = RED_FILL
        main_sheet.cell(row=7, column=144).fill = RED_FILL
        main_sheet.cell(row=7, column=145).fill = RED_FILL
        main_sheet.cell(row=7, column=146).fill = RED_FILL
        main_sheet.cell(row=7, column=147).fill = RED_FILL

    elif diag == '01010':
        sheet.cell(row=_row, column=_column).value = 'No Error'

    else:
        sheet.cell(row=_row,
                   column=_column).value = 'Invalid!'
        sheet.cell(row=_row, column=_column).fill = YELLOW_FILL
        sheet.cell(row=1, column=_column).fill = YELLOW_FILL
        main_sheet.cell(row=8, column=144).fill = YELLOW_FILL
        main_sheet.cell(row=8, column=145).fill = YELLOW_FILL
        main_sheet.cell(row=8, column=146).fill = YELLOW_FILL
        main_sheet.cell(row=8, column=147).fill = YELLOW_FILL



def translate_tle(msb, lsb):
    '''
    TLE's lexicon
    '''

    meaning = 'No Problem'

    if msb == 0 and lsb == 1:
        meaning = 'Not Available'
    elif msb == 1 and lsb == 0:
        meaning = 'Open Load'
    elif msb == 1 and lsb == 1:
        meaning = 'Short to GND'

    return meaning



def translate_tle_lowside(msb, lsb):
    '''
    TLE's lexicon (Low side)
    '''

    meaning = 'No Problem'

    if msb == 0 and lsb == 1:
        meaning = 'Short to Bat OCT'
    elif msb == 1 and lsb == 0:
        meaning = 'Open Load Off'
    elif msb == 1 and lsb == 1:
        meaning = 'Short to GND Off'

    return meaning



def translate_tle_halfbridge_lowside(msb, lsb):
    '''
    TLE's lexicon (High side)
    '''

    meaning = 'No Problem'

    if msb == 0 and lsb == 1:
        meaning = 'Not Available'
    elif msb == 1 and lsb == 0:
        meaning = 'Open Load Off'
    elif msb == 1 and lsb == 1:
        meaning = 'Short to GND Off'

    return meaning



def translate_tle_halfbridge_highside(msb, lsb):
    '''
    TLE's lexicon (High side)
    '''

    meaning = 'No Problem'

    if msb == 0 and lsb == 1:
        meaning = 'Open Load Off'
    elif msb == 1 and lsb == 0:
        meaning = 'Not Available'
    elif msb == 1 and lsb == 1:
        meaning = 'Short to Bat Off'

    return meaning



def translate_tle_pushpull(msb, lsb):
    '''
    TLE's lexicon (Push-pull)
    '''

    meaning = 'No Problem'

    if msb == 0 and lsb == 1:
        meaning = 'Short to Bat'
    elif msb == 1 and lsb == 0:
        meaning = 'Open Load Off'
    elif msb == 1 and lsb == 1:
        meaning = 'Short to GND Off'

    return meaning



def translate_etc(msb, lsb):
    '''
    ETC's lexicon
    '''

    if msb == 0 and lsb == 0:
        meaning = 'Over Current'
    elif msb == 1 and lsb == 1:
        meaning = 'Load Short!'
    else:
        meaning = 'No Problem'

    return meaning



def reading_constant_issued_values():
    '''This function creates a dictionary.'''
    parameters_dict = parameters_function()
    parameters = tuple(parameters_dict.keys())
    columns = tuple([x for x in range(1, 127) if x % 3 == 0])

    constant_values = []
    for x in columns:
        constant_values.append(main_sheet.cell(row=11,
                                               column=x - 1).value)
    constant_values = tuple(constant_values)
    constant_values_dict = dict(zip(parameters, constant_values))
    return constant_values_dict



def set_border(work_sheet, cell_range):
    '''
    Setting border for specified cells
    '''
    thin = Side(border_style="thin", color="000000")
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin,
                                 left=thin,
                                 right=thin,
                                 bottom=thin)



def file_loader():
    '''Searching and loading the files in the dictionary.'''
    filenames = []
    entries = os.scandir('TestReports')
    for entry in entries:
        forged = entry.name
        forged = 'TestReports\\' + forged
        filenames.append(forged)
    return filenames



def make_bold(work_sheet, cell_range):
    '''
    Making cells bold
    '''
    bold_font = Font(size=12, bold=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.font = bold_font



def centr_align(work_sheet, cell_range):
    '''
    Making cells center aligned
    '''
    center_alignment = Alignment(horizontal='center',
                      vertical='center',
                      wrapText=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.alignment = center_alignment



def make_blue(work_sheet, cell_range):
    '''
    Making cell text blue
    '''
    my_blue_font = Font(color=colors.COLOR_INDEX[12],
                        size=12,
                        bold=False)

    for row in work_sheet[cell_range]:
        for cell in row:
            cell.font = my_blue_font



def emi_frequencies_added_sp1(sheet):
    '''
    This function adds EMI frequencies to the specific sheet

    Argument: The specified sheet
    return  : Nothing!
    '''
    for x in range(1, EXCEL_VERT_SCOPE - 8):
        interval = \
        (int(main_sheet.cell(row=x + 9, column=1).value[0]) * 36000 + \
         int(main_sheet.cell(row=x + 9, column=1).value[1]) * 3600 + \
         int(main_sheet.cell(row=x + 9, column=1).value[3]) * 600 + \
         int(main_sheet.cell(row=x + 9, column=1).value[4]) * 60 + \
         int(main_sheet.cell(row=x + 9, column=1).value[6]) * 10 + \
         int(main_sheet.cell(row=x + 9, column=1).value[7]) + \
         int(main_sheet.cell(row=x + 9, column=1).value[9]) / 10 + \
         int(main_sheet.cell(row=x + 9, column=1).value[10]) / 100 + \
         int(main_sheet.cell(row=x + 9, column=1).value[11]) / 1000 - \
         int(main_sheet.cell(row=10, column=1).value[0]) * 36000 - \
         int(main_sheet.cell(row=10, column=1).value[1]) * 3600 - \
         int(main_sheet.cell(row=10, column=1).value[3]) * 600 - \
         int(main_sheet.cell(row=10, column=1).value[4]) * 60 - \
         int(main_sheet.cell(row=10, column=1).value[6]) * 10 - \
         int(main_sheet.cell(row=10, column=1).value[7]) - \
         int(main_sheet.cell(row=10, column=1).value[9]) / 10 - \
         int(main_sheet.cell(row=10, column=1).value[10]) / 100 - \
         int(main_sheet.cell(row=10, column=1).value[11]) / 1000)

        if interval < 60:
            sheet.cell(row=x + 1, column=9).value = \
            math.ceil(interval / 2)
        elif interval > 60:
            sheet.cell(row=x + 1, column=9).value = \
            2 * math.ceil((interval - 30) / 2)



def binary_string_to_decimal(input_string):
    '''
    This function outputs decimal, from binary string.

    Argument: The binary string
    return  : Decimal value
    '''
    decimal_value = 0
    length = len(input_string)
    for i in range(length):
        decimal_value += int(input_string[i]) * 2 ** (length - i - 1)
    return decimal_value



def convert_digitl_to_analog(digital_value, reference, resolution):
    '''
    Digital to analog conversion.

    Argument: Digital value, Reference voltage, Resolution
    return  : Analog value
    '''
    analog_value = round(digital_value * reference / 2 ** resolution, 3)
    return analog_value



def frequency_output_fluctuations(_column):
    '''
    Yellows fluctuations
    '''
    present_sheet = workbook['Frequency Outputs']
    for x in range(5, EXCEL_VERT_SCOPE - 8):
        if present_sheet.cell(row=x + 0, column=_column).value is not None and \
        present_sheet.cell(row=x + 1, column=_column).value is not None and \
        present_sheet.cell(row=x + 0, column=_column).value[:5] != \
        present_sheet.cell(row=x + 1, column=_column).value[:5] or \
        present_sheet.cell(row=x + 0, column=_column).value[-5:] != \
        present_sheet.cell(row=x + 1, column=_column).value[-5:]:
            present_sheet.cell(row=x + 1, column=_column).fill = YELLOW_FILL
            present_sheet.cell(row=2, column=_column).fill = YELLOW_FILL



def out_freq_service_pack(ugly_string, sheet, _column, _row):
    '''
    Frequency outputs service pack.

    Argument: Some ugly string
    return  : A nice informative string
    Also it does some analyzing stuff!
    '''

    OUT_FREQ_TOLERANCE = 10.0

    the_seperator = 0
    for x in range(len(ugly_string)):
        if ugly_string[x] == ' ':
            the_seperator = x

    frequency = ugly_string[:the_seperator]
    dutycycle = 100 - float(ugly_string[the_seperator + 1:])

    frequency = round(float(frequency), 1)
    dutycycle = round(float(dutycycle), 1)

    nominal_frequency = sheet.cell(row=3, column=_column).value
    nominal_dutycycle = sheet.cell(row=4, column=_column).value

    nominal_frequency = float(nominal_frequency)
    nominal_dutycycle = float(nominal_dutycycle)

    if nominal_frequency != 0 and \
    float(abs(frequency - nominal_frequency) / nominal_frequency * 100) > \
    OUT_FREQ_TOLERANCE:
            sheet.cell(row=_row, column=_column).fill = RED_FILL
            for x in range(174, 194):
                main_sheet.cell(row=7, column=x).fill = RED_FILL

    if nominal_dutycycle != 0 and \
    float(abs(dutycycle - nominal_dutycycle) / nominal_dutycycle * 100) > \
    OUT_FREQ_TOLERANCE:
            sheet.cell(row=_row, column=_column).fill = RED_FILL
            sheet.cell(row=1, column=_column).fill = RED_FILL
            for x in range(174, 194):
                main_sheet.cell(row=7, column=x).fill = RED_FILL

    result = '{}Hz  @  {}%'.format(frequency, dutycycle)
    return result



#===============================================================================
# Main Function
#===============================================================================
#---------------------------------------------------------------------- Analyzer
Analyzer_Remote_SVN_Address = 'https://192.168.5.65/svn/PECMTPE2\
/Development/Hardware/Validation/Tool/ECMT_A5_Validation/ReportCheckingApp.py'

Analyzer_Local_SVN_Address = 'ReportCheckingApp.py'

Analyzer_SVN_Class = svn.local.LocalClient(Analyzer_Local_SVN_Address)
Analyzer_SVN_Info = Analyzer_SVN_Class.info()
# print(Analyzer_SVN_Info)

Analyzer_SVN_Revision = str(Analyzer_SVN_Info['commit_revision'])
Analyzer_SVN_Committer = str(Analyzer_SVN_Info['commit_author'])
Analyzer_SVN_CommitDate = str(Analyzer_SVN_Info['commit_date'])[:-13] + ' (GMT)'

#--------------------------------------------------------------------- Test Suit
Tester_Remote_SVN_Address = 'https://192.168.5.65/svn/PECMTPE2\
/Development/Hardware/Validation/Tool/ECMT_A5_Validation\
/ECMT2_A5_GeneralManualTestSuite.robot'

Tester_Local_SVN_Address = 'ECMT2_A5_GeneralManualTestSuite.robot'

Tester_SVN_Class = svn.local.LocalClient(Tester_Local_SVN_Address)
Tester_SVN_Info = Tester_SVN_Class.info()

Tester_SVN_Revision = str(Tester_SVN_Info['commit_revision'])
Tester_SVN_Committer = str(Tester_SVN_Info['commit_author'])
Tester_SVN_CommitDate = str(Tester_SVN_Info['commit_date'])[:-13] + ' (GMT)'

#--------------------------------- Loading Files and Finding Max. Number of Rows
all_available_files = file_loader()
for file in range(len(all_available_files)):
    input_file_name = all_available_files[file]
    if '.xlsx' in input_file_name and \
       '[' not in input_file_name and \
       '$' not in input_file_name:

        print('File Processing  =>   {}'.format(input_file_name))

        output_file_name = input_file_name[:12] + \
        '[Analyzed] ' + input_file_name[12:]

        workbook = load_workbook(input_file_name,
                                 read_only=False,
                                 data_only=True)

        main_sheet = workbook[workbook.sheetnames[0]]

        main_sheet.cell(row=1, column=4).value = 'SVN Information'
        main_sheet.cell(row=2, column=4).value = 'Commit Revision'
        main_sheet.cell(row=3, column=4).value = 'Commit Author'
        main_sheet.cell(row=4, column=4).value = 'Commit Date'

        main_sheet.cell(row=1, column=6).value = 'Analyzer'
        main_sheet.cell(row=1, column=9).value = 'Test Suit'

        main_sheet.cell(row=2, column=6).value = Analyzer_SVN_Revision
        main_sheet.cell(row=3, column=6).value = Analyzer_SVN_Committer
        main_sheet.cell(row=4, column=6).value = Analyzer_SVN_CommitDate

        main_sheet.cell(row=2, column=9).value = Tester_SVN_Revision
        main_sheet.cell(row=3, column=9).value = Tester_SVN_Committer
        main_sheet.cell(row=4, column=9).value = Tester_SVN_CommitDate

        main_sheet.merge_cells('D1:E1')
        main_sheet.merge_cells('D2:E2')
        main_sheet.merge_cells('D3:E3')

        main_sheet.merge_cells('F1:H1')
        main_sheet.merge_cells('F2:H2')
        main_sheet.merge_cells('F3:H3')
        main_sheet.merge_cells('F4:H4')

        main_sheet.merge_cells('I1:K1')
        main_sheet.merge_cells('I2:K2')
        main_sheet.merge_cells('I3:K3')
        main_sheet.merge_cells('I4:K4')

        make_bold(main_sheet, 'D1:K1')
        make_bold(main_sheet, 'D1:D4')
        centr_align(main_sheet, 'F1:K1')

        EXCEL_VERT_SCOPE = len(main_sheet['A'])
        EXCEL_HRZN_SCOPE = len(main_sheet['8']) + 1

        #----------------------------------------------- Weeding Out Faulty Ones
        if EXCEL_VERT_SCOPE < 10:
            print("Invalid File; The file is Too Short!\n\n")
            continue

        if EXCEL_HRZN_SCOPE < 203:
            print("Invalid File; Some Parameters are not captured!\n\n")
            continue

        main_sheet.insert_cols(74)
        main_sheet.insert_cols(75)
        main_sheet.insert_cols(76)

        for x in range(8, 10):
            main_sheet.cell(row=x, column=74).value = \
            main_sheet.cell(row=x, column=203).value

            main_sheet.cell(row=x, column=75).value = \
            main_sheet.cell(row=x, column=204).value

            main_sheet.cell(row=x, column=76).value = \
            main_sheet.cell(row=x, column=205).value

        for x in range(10, EXCEL_VERT_SCOPE + 1):
            main_sheet.cell(row=x, column=74).value = \
            float(main_sheet.cell(row=x, column=203).value)

            main_sheet.cell(row=x, column=75).value = \
            float(main_sheet.cell(row=x, column=204).value)

            main_sheet.cell(row=x, column=76).value = \
            float(main_sheet.cell(row=x, column=205).value)

        parameters_dict = parameters_function()
        parameters = list(parameters_dict.keys())

        #=======================================================================
        # Main Page Coloring
        #=======================================================================
        #-------------------------------------------------------------- Coloring
        start_time = time.time()
        for _column in [x for x in range(1, 127) if x % 6 == 0]:
            for cell in main_sheet.iter_rows(min_row=8,
                                             max_row=EXCEL_VERT_SCOPE,
                                             min_col=_column - 1,
                                             max_col=_column + 1):
                for j in range(len(cell)):
                    cell[j].fill = PEACH_FILL

        for _column in [x for x in range(1, 127) if x % 3 == 0 and x % 2 != 0]:
            for cell in main_sheet.iter_rows(min_row=8,
                                             max_row=EXCEL_VERT_SCOPE,
                                             min_col=_column - 1,
                                             max_col=_column + 1):
                for j in range(len(cell)):
                    cell[j].fill = BLUE_FILL

        for _column in [x for x in range(128, EXCEL_HRZN_SCOPE) if x % 2 == 0]:
            for cell in main_sheet.iter_rows(min_row=8,
                                             max_row=EXCEL_VERT_SCOPE,
                                             min_col=_column,
                                             max_col=_column):
                for j in range(len(cell)):
                    cell[j].fill = LGREY_FILL

        for _column in [x for x in range(128, EXCEL_HRZN_SCOPE) if x % 2 != 0]:
            for cell in main_sheet.iter_rows(min_row=8,
                                             max_row=EXCEL_VERT_SCOPE,
                                             min_col=_column,
                                             max_col=_column):
                for j in range(len(cell)):
                    cell[j].fill = GREY_FILL

        #=======================================================================
        # CAN Failure Diagnosis
        #=======================================================================
        CAN_failure_times = []

        vital_flag = 0
        IGK_value = 13.5

        for _row in range(10, EXCEL_VERT_SCOPE):
            if main_sheet.cell(row=_row, column=2).value != 'CAN Failure!':
                IGK_value = float(main_sheet.cell(row=_row, column=2).value)
                main_sheet.cell(row=11, column=2).value = \
                float(main_sheet.cell(row=_row, column=2).value)
                main_sheet.cell(row=10, column=2).value = \
                float(main_sheet.cell(row=_row, column=2).value)
                vital_flag = 1
                break

        if vital_flag == 0:
            print('This file is initially plagued by CAN failure!\n\n')
            continue

        constant_issued_values = reading_constant_issued_values()

        for _row in range(10, EXCEL_VERT_SCOPE):
                if main_sheet.cell(row=1, column=2).value == 'CAN Failure!':
                    CAN_failure_times.append(main_sheet.cell(row=_row,
                                                             column=1).value)
                    main_sheet.cell(row=9, column=1).fill = YELLOW_FILL

        for _row in range(1, EXCEL_VERT_SCOPE - 1):
            if main_sheet.cell(row=_row, column=2).value == 'CAN Failure!':
                for cell in main_sheet.iter_rows(min_row=_row,
                                             max_row=_row,
                                             min_col=1,
                                             max_col=EXCEL_HRZN_SCOPE - 1):
                    for j in range(len(cell)):
                        cell[j].fill = VIOLET_FILL

        #=======================================================================
        # LAN Failure Diagnosis
        #=======================================================================
        LAN_failure_times = []

        for _row in range(10, EXCEL_VERT_SCOPE - 1):
            upper_time = int(main_sheet.cell(row=_row,
                                             column=1).value[6:8]) + \
            int(main_sheet.cell(row=_row, column=1).value[3:5]) * 60 + \
            int(main_sheet.cell(row=_row, column=1).value[:2]) * 3600

            lower_time = int(main_sheet.cell(row=_row + 1,
                                             column=1).value[6:8]) + \
            int(main_sheet.cell(row=_row + 1, column=1).value[3:5]) * 60 + \
            int(main_sheet.cell(row=_row + 1, column=1).value[:2]) * 3600

            time_step = abs(lower_time - upper_time)
            if time_step > 4:
                LAN_failure_times.append(main_sheet.cell(row=_row,
                                                         column=1).value)
                LAN_failure_times.append(main_sheet.cell(row=_row + 1,
                                                         column=1).value)
                for j in range(1, EXCEL_HRZN_SCOPE):
                    main_sheet.cell(row=_row, column=j).fill = BROWN_FILL
                    main_sheet.cell(row=_row + 1, column=j).fill = BROWN_FILL
                    for cell in main_sheet.iter_rows(min_row=_row,
                                             max_row=_row,
                                             min_col=1,
                                             max_col=EXCEL_HRZN_SCOPE - 1):
                        for j in range(len(cell)):
                            cell[j].fill = BROWN_FILL

                main_sheet.cell(row=9, column=1).fill = YELLOW_FILL
                main_sheet.cell(row=_row, column=2).value = 'LAN Failure!'

        #=======================================================================
        # Creating Excel Sheets
        #=======================================================================
        '''
        This is is one of the most important features of our script.
        This process, deletes old file's sheets first.
        (Except the main sheet, of course.)'''

        for sheet in workbook.sheetnames:
            if sheet != workbook.sheetnames[0]:
                workbook.remove(workbook[sheet])

        ''' We could chose to have one sheet
        per any frequency or digital input, but we didn't!
        '''
        for param in parameters[:ANALOG_INPUTS + 0 + 0]:
            workbook.create_sheet(param)
            present_sheet = workbook[param]
            for column in Excel_std_column_names[:8]:
                present_sheet.column_dimensions[column].width = 21
            present_sheet.column_dimensions[Excel_std_column_names[8]].width = 4
            for column in Excel_std_column_names[9:15]:
                present_sheet.column_dimensions[column].width = 16

        workbook.create_sheet('Digital Inputs')
        # workbook.create_sheet('CAM')
        workbook.create_sheet('WSS')
        workbook.create_sheet('ALT_MON')
        # workbook.create_sheet('CAM_EX')
        workbook.create_sheet('CRANK')
        workbook.create_sheet('TLE8888 Registers')
        workbook.create_sheet('ETC Registers')
        workbook.create_sheet('KP254 Values')
        workbook.create_sheet('Temperature Values')
        workbook.create_sheet('Reset Registers')
        workbook.create_sheet('Monitoring Unit')
        workbook.create_sheet('Frequency Outputs')

        #----------------------------------------------- Format of Main function
        '''

        #=======================================================================
        # Parameter
        #=======================================================================
        #------------------------------------------------------------ Processing
        #----------------------------------------------------------------- Chart
        #----------------------------------------------- Coloring and Statistics

        '''
        #=======================================================================
        # Analog Inputs
        #=======================================================================
        #------------------------------------------------------------ Processing

        for param in parameters[:ANALOG_INPUTS]:
            present_sheet = workbook[param]
            data = analog_service_pack(param)

            for _column in range(len(data[0])):
                present_sheet.cell(row=1,
                                   column=_column + 1).value = data[0][_column]
                present_sheet.cell(row=2,
                                   column=_column + 1).value = data[1][_column]

            for x in range(1, EXCEL_VERT_SCOPE - 7):
                present_sheet.cell(row=x + 0, column=10).value = \
                main_sheet.cell(row=x + 8, column=1).value

                if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                    present_sheet.cell(row=x + 0, column=10).fill = BROWN_FILL
                    present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

                if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                    present_sheet.cell(row=x + 0, column=10).fill = VIOLET_FILL
                    present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

                if main_sheet.cell(row=10,
                                   column=parameters_dict[param] - 1)\
                                   .value != '??':

                    present_sheet.cell(row=1, column=13).value = 'Issued Value'
                    present_sheet.cell(row=1, column=14).value = '+5% Error'
                    present_sheet.cell(row=1, column=15).value = '-5% Error'

                    present_sheet.cell(row=x + 0, column=13).value = \
                    main_sheet.cell(row=x + 8,
                                    column=parameters_dict[param] - 1).value

                    present_sheet.cell(row=x + 0, column=14).value = \
                    float(main_sheet.cell(row=10,
                                          column=parameters_dict[param] - 1)\
                                          .value) * (1 + TOLERANCE)

                    present_sheet.cell(row=x + 0, column=15).value = \
                    float(main_sheet.cell(row=10,
                                          column=parameters_dict[param] - 1)\
                                          .value) * (1 - TOLERANCE)

                present_sheet.cell(row=x + 0, column=11).value = \
                main_sheet.cell(row=x + 8, column=parameters_dict[param]).value

                present_sheet.cell(row=x + 0, column=12).value = \
                main_sheet.cell(row=x + 8, column=parameters_dict[param] + 1)\
                .value

            if main_sheet.cell(row=10,
                               column=parameters_dict[param] - 1).value != '??':
                for x in range(5, EXCEL_VERT_SCOPE - 7):
                    if float(present_sheet.cell(row=x, column=12).value)\
                     > float(present_sheet.cell(row=x, column=14).value)\
                     or\
                       float(present_sheet.cell(row=x, column=12).value)\
                     < float(present_sheet.cell(row=x, column=15).value):
                        present_sheet.cell(row=x, column=12).fill = RED_FILL
                        present_sheet.cell(row=1, column=12).fill = RED_FILL
                        main_sheet.cell(row=7,
                                        column=parameters_dict[param]).fill\
 = RED_FILL

                    if float(present_sheet.cell(row=x, column=11).value)\
                     < float(present_sheet.cell(row=x, column=15).value)\
                     or\
                       float(present_sheet.cell(row=x, column=11).value)\
                     > float(present_sheet.cell(row=x, column=14).value):
                        present_sheet.cell(row=x, column=11).fill = RED_FILL
                        present_sheet.cell(row=1, column=11).fill = RED_FILL
                        main_sheet.cell(row=7,
                                        column=parameters_dict[param]).fill\
 = RED_FILL

            centr_align(present_sheet, 'A1:O{}'.format(EXCEL_VERT_SCOPE))
            make_bold(present_sheet, 'A1:O1')

        #----------------------------------------------------------------- Chart
            chart = LineChart()
            if main_sheet.cell(row=10,
                               column=parameters_dict[param] - 1).value != '??':
                data = Reference(worksheet=present_sheet,
                                  min_row=1,
                                  max_row=EXCEL_VERT_SCOPE,
                                  min_col=11,
                                  max_col=15)
            else:
                data = Reference(worksheet=present_sheet,
                                  min_row=1,
                                  max_row=EXCEL_VERT_SCOPE,
                                  min_col=11,
                                  max_col=12)

            chart.add_data(data, titles_from_data=True)

            cats = Reference(worksheet=present_sheet,
                              min_row=2,
                              max_row=EXCEL_VERT_SCOPE,
                              min_col=10,
                              max_col=10)
            chart.set_categories(cats)

            if main_sheet.cell(row=10,
                               column=parameters_dict[param] - 1).value != '??':
                botE = chart.series[3]
                topE = chart.series[4]
                topE.graphicalProperties.line.dashStyle = "sysDot"
                botE.graphicalProperties.line.dashStyle = "sysDot"

            chart.x_axis.title = "Sample Time"
            chart.y_axis.title = main_sheet.cell\
                                (row=8,
                                 column=parameters_dict[param]
                                ).value

            chart.width = CHART_WIDTH
            chart.height = CHART_HIGHT

            present_sheet.add_chart(chart, "A5")

        #----------------------------------------------- Coloring and Statistics
            present_sheet.freeze_panes = 'A2'

            if main_sheet.cell(row=10,
                               column=parameters_dict[param] - 1).value != '??':
                for cell in present_sheet.iter_rows(min_row=1,
                                                    max_row=2,
                                                    min_col=1,
                                                    max_col=8):
                    for j in range(len(cell)):
                        cell[j].fill = GREEN_FILL
                set_border(present_sheet, 'A0:H2')
            else:
                for cell in present_sheet.iter_rows(min_row=1,
                                                    max_row=2,
                                                    min_col=1,
                                                    max_col=6):
                    for j in range(len(cell)):
                        cell[j].fill = GREEN_FILL
                set_border(present_sheet, 'A0:F2')

        main_sheet.delete_cols(74, 3)

        #=======================================================================
        # Digital Inputs
        #=======================================================================
        #------------------------------------------------------------ Processing
        # print("Digital Inputs")
        present_sheet = workbook['Digital Inputs']

        for x in range(1, EXCEL_VERT_SCOPE - 6):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 7, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=2, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=2, column=1).fill = YELLOW_FILL

            present_sheet.cell(row=1, column=2).value = 'Parameter'

        y = 2
        for k in range (98 - 74):
            for x in range(1, EXCEL_VERT_SCOPE - 6):
                present_sheet.cell(row=x + 0,
                                   column=y
                                         +k
                                         +math.floor((2 - k % 3) / 2)).value = \
                main_sheet.cell(row=x + 7, column=74 + k).value
            y += math.floor((2 - k % 3) / 2)

        for k in range (200 - 197):
            for x in range(1, EXCEL_VERT_SCOPE - 6):
                present_sheet.cell(row=x + 0,
                                   column=35 + k).value = \
                main_sheet.cell(row=x + 7, column=197 + k).value

        clmns = (4, 8, 12, 20, 24, 28, 32, 36)
        rws = range(6, EXCEL_VERT_SCOPE - 6)

        for rw in rws:
            if present_sheet.cell(row=rw,
                                  column=16).value == '1':
                present_sheet.cell(row=rw,
                                   column=16).fill = GREEN_FILL
            else:
                present_sheet.cell(row=rw,
                                   column=16).fill = RED_FILL
                present_sheet.cell(row=1,
                                   column=16).fill = RED_FILL
                for k in range(75, 98):
                    main_sheet.cell(row=7,
                                    column=k).fill = RED_FILL

        for clmn in clmns:
            for rw in rws:
                if present_sheet.cell(row=rw,
                                      column=clmn - 1).value == '1' and \
                   present_sheet.cell(row=rw,
                                      column=clmn).value != '1':
                    present_sheet.cell(row=rw,
                                       column=clmn).fill = RED_FILL
                    present_sheet.cell(row=1,
                                       column=clmn).fill = RED_FILL
                    for k in range(75, 98):
                        main_sheet.cell(row=7,
                                        column=k).fill = RED_FILL

                else:
                    present_sheet.cell(row=rw, column=clmn).fill = GREEN_FILL

                if present_sheet.cell(row=rw,
                                      column=clmn - 1).value != \
                   present_sheet.cell(row=rw - 1,
                                      column=clmn - 1).value and \
                   present_sheet.cell(row=rw,
                                      column=clmn + 1).value != '1':
                    present_sheet.cell(row=rw,
                                       column=clmn + 1).fill = RED_FILL
                    present_sheet.cell(row=1,
                                       column=clmn + 1).fill = RED_FILL
                    for k in range(75, 98):
                        main_sheet.cell(row=7,
                                        column=k).fill = RED_FILL

                elif present_sheet.cell(row=rw,
                                        column=clmn + 1).value == '1':
                    present_sheet.cell(row=rw,
                                       column=clmn + 1).fill = YELLOW_FILL

        #----------------------------------------------- Coloring and Statistics
        present_sheet.freeze_panes = 'A3'

        centr_align(present_sheet, 'A1:AK{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:AK2')

        present_sheet.column_dimensions[Excel_std_column_names[0]].width = 15
        for column in Excel_std_column_names[1:32]:
            present_sheet.column_dimensions[column].width = 7

        #=======================================================================
        # Frequency Inputs
        #=======================================================================
        #------------------------------------------------------------ Processing
        # print("Frequency Inputs")
        freq_sheets = ('WSS', 'ALT_MON', 'CRANK')
        freq_clmns = ([105, 7000, 3000, '100Hz 70%'],
                      [111, 400, 3600, '250Hz 10%'],
                      [123, 83.3, 83.3, '6kHz 50%'])
        freq_sheets_dict = dict(zip(freq_sheets, freq_clmns))

        for sheet in freq_sheets:
            present_sheet = workbook[sheet]

            for x in range(1, EXCEL_VERT_SCOPE - 7):
                present_sheet.cell(row=x + 0, column=10).value = \
                main_sheet.cell(row=x + 7, column=1).value

                if main_sheet.cell(row=x + 7, column=2).value == 'LAN Failure!':
                    present_sheet.cell(row=x + 0, column=10).fill = BROWN_FILL
                    present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

                if main_sheet.cell(row=x + 7, column=2).value == 'CAN Failure!':
                    present_sheet.cell(row=x + 0, column=10).fill = VIOLET_FILL
                    present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

            set_border(present_sheet, 'A1:D2')

            present_sheet.cell(row=1,
                               column=1).value = 'Parameter'
            present_sheet.cell(row=2,
                               column=1).value = sheet

            present_sheet.cell(row=1,
                               column=2).value = 'Issued Freq.'
            present_sheet.cell(row=2,
                               column=2).value = freq_sheets_dict[sheet][3]

            present_sheet.cell(row=1,
                               column=3).value = 'Issued Low'
            present_sheet.cell(row=2,
                               column=3).value = freq_sheets_dict[sheet][2]

            present_sheet.cell(row=1,
                               column=4).value = 'Issued High'
            present_sheet.cell(row=2,
                               column=4).value = freq_sheets_dict[sheet][1]

            present_sheet.cell(row=1,
                               column=13).value = 'Issued'
            present_sheet.cell(row=2,
                               column=13).value = 'Issued High'

            present_sheet.cell(row=1,
                               column=11).value = 'High Value'
            present_sheet.cell(row=2,
                               column=11).value = 'Min'

            present_sheet.cell(row=1,
                               column=12).value = 'High Value'
            present_sheet.cell(row=2,
                               column=12).value = 'Max'

            present_sheet.cell(row=1,
                               column=14).value = 'High Value'
            present_sheet.cell(row=2,
                               column=14).value = '+5%'

            present_sheet.cell(row=1,
                               column=15).value = 'High Value'
            present_sheet.cell(row=2,
                               column=15).value = '-5%'

            present_sheet.cell(row=1,
                               column=18).value = 'Issued'
            present_sheet.cell(row=2,
                               column=18).value = 'Issued Low'

            present_sheet.cell(row=1,
                               column=16).value = 'Low Value'
            present_sheet.cell(row=2,
                               column=16).value = 'Min'

            present_sheet.cell(row=1,
                               column=17).value = 'Low Value'
            present_sheet.cell(row=2,
                               column=17).value = 'Max'

            present_sheet.cell(row=1,
                               column=19).value = 'Low Value'
            present_sheet.cell(row=2,
                               column=19).value = '+5%'

            present_sheet.cell(row=1,
                               column=20).value = 'Low Value'
            present_sheet.cell(row=2,
                               column=20).value = '-5%'

            for x in range(3, EXCEL_VERT_SCOPE - 7):
                present_sheet.cell(row=x + 0, column=13).value = \
                int(present_sheet.cell(row=2, column=4).value)

                present_sheet.cell(row=x + 0, column=11).value = \
                main_sheet.cell(row=x + 7,
                                column=freq_sheets_dict[sheet][0]).value

                present_sheet.cell(row=x + 0,
                                   column=12).value = \
                main_sheet.cell(row=x + 7,
                                column=freq_sheets_dict[sheet][0] + 1).value

                present_sheet.cell(row=x + 0, column=14).value = \
                float(present_sheet.cell(row=2,
                                         column=4).value) * 1.05
                present_sheet.cell(row=x + 0, column=15).value = \
                float(present_sheet.cell(row=2,
                                         column=4).value) * 0.95

                present_sheet.cell(row=x + 0, column=18).value = \
                int(present_sheet.cell(row=2, column=3).value)

                present_sheet.cell(row=x + 0, column=16).value = \
                main_sheet.cell(row=x + 7,
                                column=freq_sheets_dict[sheet][0] + 3).value
                present_sheet.cell(row=x + 0, column=17).value = \
                main_sheet.cell(row=x + 7,
                                column=freq_sheets_dict[sheet][0] + 4).value

                present_sheet.cell(row=x + 0, column=19).value = \
                float(present_sheet.cell(row=2,
                                         column=3).value) * 1.05
                present_sheet.cell(row=x + 0, column=20).value = \
                float(present_sheet.cell(row=2,
                                         column=3).value) * 0.95

            for x in range(6, EXCEL_VERT_SCOPE - 8):
                if float(present_sheet.cell(row=x + 0, column=11).value) < \
                   float(present_sheet.cell(row=x + 0, column=15).value)\
                   or\
                   float(present_sheet.cell(row=x + 0, column=11).value) > \
                   float(present_sheet.cell(row=x + 0, column=14).value):
                    present_sheet.cell(row=x + 0, column=11).fill = RED_FILL
                    present_sheet.cell(row=1, column=11).fill = RED_FILL
                    if sheet == 'WSS':
                        for k in range(95, 110):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'ALT_MON':
                        for k in range(110, 115):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'CRANK':
                        for k in range(122, 128):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                # else:
                #     present_sheet.cell(row=x + 0, column=11).fill = GREEN_FILL

                if float(present_sheet.cell(row=x + 0, column=12).value) > \
                   float(present_sheet.cell(row=x + 0, column=14).value)\
                   or\
                   float(present_sheet.cell(row=x + 0, column=12).value) < \
                   float(present_sheet.cell(row=x + 0, column=15).value):
                    present_sheet.cell(row=x + 0, column=12).fill = RED_FILL
                    present_sheet.cell(row=1, column=12).fill = RED_FILL
                    if sheet == 'WSS':
                        for k in range(95, 110):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'ALT_MON':
                        for k in range(110, 115):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'CRANK':
                        for k in range(122, 128):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL

                # else:
                #     present_sheet.cell(row=x + 0, column=12).fill = GREEN_FILL

                if float(present_sheet.cell(row=x + 0, column=16).value) < \
                   float(present_sheet.cell(row=x + 0, column=20).value)\
                   or\
                   float(present_sheet.cell(row=x + 0, column=16).value) > \
                   float(present_sheet.cell(row=x + 0, column=19).value):
                    present_sheet.cell(row=x + 0, column=16).fill = RED_FILL
                    present_sheet.cell(row=1, column=16).fill = RED_FILL

                    if sheet == 'WSS':
                        for k in range(95, 110):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'ALT_MON':
                        for k in range(110, 115):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'CRANK':
                        for k in range(122, 128):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                # else:
                #     present_sheet.cell(row=x + 0, column=16).fill = GREEN_FILL

                if float(present_sheet.cell(row=x + 0, column=17).value) > \
                   float(present_sheet.cell(row=x + 0, column=19).value)\
                   or\
                   float(present_sheet.cell(row=x + 0, column=17).value) < \
                   float(present_sheet.cell(row=x + 0, column=20).value):
                    present_sheet.cell(row=x + 0, column=17).fill = RED_FILL
                    present_sheet.cell(row=1, column=17).fill = RED_FILL

                    if sheet == 'WSS':
                        for k in range(95, 110):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'ALT_MON':
                        for k in range(110, 115):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                    elif sheet == 'CRANK':
                        for k in range(122, 128):
                            main_sheet.cell(row=7, column=k).fill = RED_FILL
                # else:
                #     present_sheet.cell(row=x + 0, column=17).fill = GREEN_FILL

            centr_align(present_sheet, 'A1:T{}'.format(EXCEL_VERT_SCOPE))
            make_bold(present_sheet, 'J1:T2')
            make_bold(present_sheet, 'A1:I1')

            for column in Excel_std_column_names[0:9]:
                present_sheet.column_dimensions[column].width = 18

            for column in Excel_std_column_names[9:10]:
                present_sheet.column_dimensions[column].width = 18

            for column in Excel_std_column_names[10:20]:
                present_sheet.column_dimensions[column].width = 12

            present_sheet.freeze_panes = 'A2'

        #----------------------------------------------------------------- Chart
            chart = LineChart()

            data = Reference(worksheet=present_sheet,
                                  min_row=2,
                                  max_row=EXCEL_VERT_SCOPE - 10,
                                  min_col=11,
                                  max_col=15)

            chart.add_data(data, titles_from_data=True)

            cats = Reference(worksheet=present_sheet,
                              min_row=2,
                              max_row=EXCEL_VERT_SCOPE,
                              min_col=10,
                              max_col=10)

            chart.set_categories(cats)

            chart.x_axis.title = "Sample Time"
            chart.y_axis.title = sheet

            botE = chart.series[3]
            topE = chart.series[4]
            topE.graphicalProperties.line.dashStyle = "sysDot"
            botE.graphicalProperties.line.dashStyle = "sysDot"

            chart.width = CHART_WIDTH
            chart.height = CHART_HIGHT

            present_sheet.add_chart(chart, "A5")

            chart2 = LineChart()

            data = Reference(worksheet=present_sheet,
                                  min_row=2,
                                  max_row=EXCEL_VERT_SCOPE - 10,
                                  min_col=16,
                                  max_col=20)

            chart2.add_data(data, titles_from_data=True)

            cats = Reference(worksheet=present_sheet,
                              min_row=2,
                              max_row=EXCEL_VERT_SCOPE,
                              min_col=10,
                              max_col=10)

            chart2.set_categories(cats)

            chart2.x_axis.title = "Sample Time"
            chart2.y_axis.title = sheet

            botE = chart2.series[3]
            topE = chart2.series[4]
            topE.graphicalProperties.line.dashStyle = "sysDot"
            botE.graphicalProperties.line.dashStyle = "sysDot"

            chart2.width = CHART_WIDTH
            chart2.height = CHART_HIGHT

            present_sheet.add_chart(chart2, "A40")

        #=======================================================================
        # TLE Registries
        #=======================================================================
        # print("TLE Registries")

        '''
        [TLE8888 Diagnosis]
        http://192.168.5.62:8090/display/ECMT2/ID-022%3A+Drive+TLE8888

        [TLE8888 Report Request]
        http://192.168.5.62:8090/display/ECMT2/0x53+-+SBC+TLE8888+Report+Request
        '''

        #------------------------------------------------------------ Processing
        present_sheet = workbook['TLE8888 Registers']

        present_sheet.cell(row=1, column=2).value = 'IgnDiag'
        present_sheet.cell(row=1, column=3).value = 'BriDiag1'
        present_sheet.cell(row=1, column=4).value = 'BriDiag0'
        present_sheet.cell(row=1, column=5).value = 'OutDiag4'
        present_sheet.cell(row=1, column=6).value = 'OutDiag3'
        present_sheet.cell(row=1, column=7).value = 'OutDiag2'
        present_sheet.cell(row=1, column=8).value = 'OutDiag1'
        present_sheet.cell(row=1, column=9).value = 'OutDiag0'

        present_sheet.cell(row=1, column=10).value = 'IGN 1'
        present_sheet.cell(row=1, column=11).value = 'IGN 2'
        present_sheet.cell(row=1, column=12).value = 'IGN 3'
        present_sheet.cell(row=1, column=13).value = 'IGN 4'

        present_sheet.cell(row=1, column=14).value = 'IV1_CYL1'
        present_sheet.cell(row=1, column=15).value = 'IV2_CYL3'
        present_sheet.cell(row=1, column=16).value = 'IV3_CYL4'
        present_sheet.cell(row=1, column=17).value = 'IV4_CYL2'

        present_sheet.cell(row=1, column=18).value = 'DO_IVVTPWM'
        present_sheet.cell(row=1, column=19).value = 'LSHPWM_UP'
        present_sheet.cell(row=1, column=20).value = 'LSHPWM_DN'
        present_sheet.cell(row=1, column=21).value = 'DO_WGPWM'

        present_sheet.cell(row=1, column=22).value = 'CPPWM'
        present_sheet.cell(row=1, column=23).value = 'DO_EVVTPWM'
        present_sheet.cell(row=1, column=24).value = 'RLY_EFP'
        present_sheet.cell(row=1, column=25).value = 'RLY_ACCOUT'

        present_sheet.cell(row=1, column=26).value = 'RLY_FAN_HI'
        present_sheet.cell(row=1, column=27).value = 'RLY_STST'
        present_sheet.cell(row=1, column=28).value = 'RLY_VAC_PuMP'
        present_sheet.cell(row=1, column=29).value = 'HOT_LAMP'

        present_sheet.cell(row=1, column=30).value = 'MIL'
        present_sheet.cell(row=1, column=31).value = 'RLY_START'
        present_sheet.cell(row=1, column=32).value = 'RLY_FAN_LOW'
        present_sheet.cell(row=1, column=33).value = 'ALT_CMD'

        present_sheet.cell(row=1, column=34).value = 'ELE_THERMOSTAT'
        present_sheet.cell(row=1, column=35).value = 'ELE_WATER_PUMP'
        present_sheet.cell(row=1, column=36).value = 'DO_ELE_OIL_PUMP'
        present_sheet.cell(row=1, column=37).value = 'DO_RCL'
        present_sheet.cell(row=1, column=38).value = 'Diag0_OCT'

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=2).value = \
            main_sheet.cell(row=x + 9, column=128).value

            present_sheet.cell(row=x + 1, column=3).value = \
            main_sheet.cell(row=x + 9, column=129).value

            present_sheet.cell(row=x + 1, column=4).value = \
            main_sheet.cell(row=x + 9, column=130).value

            present_sheet.cell(row=x + 1, column=5).value = \
            main_sheet.cell(row=x + 9, column=131).value

            present_sheet.cell(row=x + 1, column=6).value = \
            main_sheet.cell(row=x + 9, column=132).value

            present_sheet.cell(row=x + 1, column=7).value = \
            main_sheet.cell(row=x + 9, column=133).value

            present_sheet.cell(row=x + 1, column=8).value = \
            main_sheet.cell(row=x + 9, column=134).value

            present_sheet.cell(row=x + 1, column=9).value = \
            main_sheet.cell(row=x + 9, column=135).value

        #--------------------------------------------------------------- IGNDIAG
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            '''
            IGN4
            MSb       : Byte [bit] = IGNDIAG [7]
            LSb       : Byte [bit] = IGNDIAG [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 6])
            present_sheet.cell(row=x + 1, column=13).value = \
            translate_tle(msb, lsb)
            if present_sheet.cell(row=x + 1, column=13).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=13).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=13).fill = RED_FILL
                present_sheet.cell(row=1, column=13).fill = RED_FILL

            '''
            IGN3
            MSb       : Byte [bit] = IGNDIAG [5]
            LSb       : Byte [bit] = IGNDIAG [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 4])
            present_sheet.cell(row=x + 1, column=12).value = \
            translate_tle(msb, lsb)
            if present_sheet.cell(row=x + 1, column=12).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=12).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=12).fill = RED_FILL
                present_sheet.cell(row=1, column=12).fill = RED_FILL

            '''
            IGN2
            MSb       : Byte [bit] = IGNDIAG [3]
            LSb       : Byte [bit] = IGNDIAG [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 2])
            present_sheet.cell(row=x + 1, column=11).value = \
            translate_tle(msb, lsb)
            if present_sheet.cell(row=x + 1, column=11).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=11).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=11).fill = RED_FILL
                present_sheet.cell(row=1, column=11).fill = RED_FILL

            '''
            IGN1
            MSb       : Byte [bit] = IGNDIAG [1]
            LSb       : Byte [bit] = IGNDIAG [0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[7 - 0])
            present_sheet.cell(row=x + 1, column=10).value = \
            translate_tle(msb, lsb)
            if present_sheet.cell(row=x + 1, column=10).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=10).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=10).fill = RED_FILL
                present_sheet.cell(row=1, column=10).fill = RED_FILL

        #--------------------------------------------------- BRIDIAG0 & BRIDIAG1
            '''
            RLY_EFP
            Pin       : 24
            MSb       : Byte [bit] = BRIDIAG0 [7]
            LSb       : Byte [bit] = BRIDIAG0 [6]
            Related to: Byte [bit] = BRIDIAG1 [3]
            Related to: Byte [bit] = BRIDIAG1 [5]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 6])
            if int(present_sheet.cell(row=x + 1, column=3).value[7 - 5]) == 1:
                present_sheet.cell(row=x + 1, column=24).value = \
                'Over Temperature'
            elif int(present_sheet.cell(row=x + 1, column=3).value[7 - 3]) == 1:
                present_sheet.cell(row=x + 1, column=24).value = \
                'Over Current'
            else:
                present_sheet.cell(row=x + 1, column=24).value = \
                translate_tle_halfbridge_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=24).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=24).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=24).fill = RED_FILL
                present_sheet.cell(row=1, column=24).fill = RED_FILL

            '''
            RLY_ACCOUT
            Pin       : 23
            MSb       : Byte [bit] = BRIDIAG0 [5]
            LSb       : Byte [bit] = BRIDIAG0 [4]
            Related to: Byte [bit] = BRIDIAG1 [2]
            Related to: Byte [bit] = BRIDIAG1 [5]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 4])
            if int(present_sheet.cell(row=x + 1, column=3).value[7 - 5]) == 1:
                present_sheet.cell(row=x + 1, column=25).value = \
                'Over Temperature'
            elif int(present_sheet.cell(row=x + 1, column=3).value[7 - 2]) == 1:
                present_sheet.cell(row=x + 1, column=25).value = \
                'Over Current'
            else:
                present_sheet.cell(row=x + 1, column=25).value = \
                translate_tle_halfbridge_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=25).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=25).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=25).fill = RED_FILL
                present_sheet.cell(row=1, column=25).fill = RED_FILL

            '''
            RLY_FAN_HIGH
            Pin       : 22
            MSb       : Byte [bit] = BRIDIAG0 [3]
            LSb       : Byte [bit] = BRIDIAG0 [2]
            Related to: Byte [bit] = BRIDIAG1 [1]
            Related to: Byte [bit] = BRIDIAG1 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 2])
            if int(present_sheet.cell(row=x + 1, column=3).value[7 - 4]) == 1:
                present_sheet.cell(row=x + 1, column=26).value = \
                'Over Temperature'
            elif int(present_sheet.cell(row=x + 1, column=3).value[7 - 1]) == 1:
                present_sheet.cell(row=x + 1, column=26).value = \
                'Over Current'
            else:
                present_sheet.cell(row=x + 1, column=26).value = \
                translate_tle_halfbridge_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=26).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=26).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=26).fill = RED_FILL
                present_sheet.cell(row=1, column=26).fill = RED_FILL

            '''
            RLY_STST
            Pin       : 21
            MSb       : Byte [bit] = BRIDIAG0 [1]
            LSb       : Byte [bit] = BRIDIAG0 [0]
            Related to: Byte [bit] = BRIDIAG1 [0]
            Related to: Byte [bit] = BRIDIAG1 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=4).value[7 - 0])
            if int(present_sheet.cell(row=x + 1, column=3).value[7 - 4]) == 1:
                present_sheet.cell(row=x + 1, column=27).value = \
                'Over Temperature'
            elif int(present_sheet.cell(row=x + 1, column=3).value[7 - 0]) == 1:
                present_sheet.cell(row=x + 1, column=27).value = \
                'Over Current'
            else:
                present_sheet.cell(row=x + 1, column=27).value = \
                translate_tle_halfbridge_highside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=27).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=27).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=27).fill = RED_FILL
                present_sheet.cell(row=1, column=27).fill = RED_FILL

            #---------------------------------------------------------- OUTDIAG4
            '''
            RLY_VAC_PUMP
            Pin       : 20
            MSb       : Byte [bit] = OUTDIAG4 [7]
            LSb       : Byte [bit] = OUTDIAG4 [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 6])
            present_sheet.cell(row=x + 1, column=28).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=28).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=28).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=28).fill = RED_FILL
                present_sheet.cell(row=1, column=28).fill = RED_FILL

            '''
            HOT_LAMP
            Pin       : 19
            MSb       : Byte [bit] = OUTDIAG4 [5]
            LSb       : Byte [bit] = OUTDIAG4 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 4])
            present_sheet.cell(row=x + 1, column=29).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=29).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=29).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=29).fill = RED_FILL
                present_sheet.cell(row=1, column=29).fill = RED_FILL

            '''
            MIL
            Pin       : 18
            MSb       : Byte [bit] = OUTDIAG4 [3]
            LSb       : Byte [bit] = OUTDIAG4 [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 2])
            present_sheet.cell(row=x + 1, column=30).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=30).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=30).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=30).fill = RED_FILL
                present_sheet.cell(row=1, column=30).fill = RED_FILL

            '''
            RLY_START
            Pin       : 17
            MSb       : Byte [bit] = OUTDIAG4 [1]
            LSb       : Byte [bit] = OUTDIAG4 [0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=5).value[7 - 0])
            present_sheet.cell(row=x + 1, column=31).value = \
            translate_tle(msb, lsb)
            if present_sheet.cell(row=x + 1, column=31).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=31).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=31).fill = RED_FILL
                present_sheet.cell(row=1, column=31).fill = RED_FILL

            #---------------------------------------------------------- OUTDIAG3
            '''
            RLY_FAN_LOW
            Pin       : 16
            MSb       : Byte[bit] = OUTDIAG3 [7]
            LSb       : Byte[bit] = OUTDIAG3 [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 6])
            present_sheet.cell(row=x + 1, column=32).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=32).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=32).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=32).fill = RED_FILL
                present_sheet.cell(row=1, column=32).fill = RED_FILL

            '''
            ALT_CMD
            Pin       : 15
            MSb       : Byte [bit] = OUTDIAG3 [5]
            LSb       : Byte [bit] = OUTDIAG3 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 4])
            present_sheet.cell(row=x + 1, column=33).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=33).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=33).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=33).fill = RED_FILL
                present_sheet.cell(row=1, column=33).fill = RED_FILL

            '''
            ELE_THERMOSTAT
            Pin       : 14
            MSb       : Byte [bit] = OUTDIAG3 [3]
            LSb       : Byte [bit] = OUTDIAG3 [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 2])
            present_sheet.cell(row=x + 1, column=34).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=34).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=34).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=34).fill = RED_FILL
                present_sheet.cell(row=1, column=34).fill = RED_FILL

            '''
            ELE_WATER_PUMP
            Pin       : 13
            MSb       : Byte [bit] = OUTDIAG3 [1]
            LSb       : Byte [bit] = OUTDIAG3 [0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=6).value[7 - 0])
            present_sheet.cell(row=x + 1, column=35).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=35).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=35).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=35).fill = RED_FILL
                present_sheet.cell(row=1, column=35).fill = RED_FILL

            #---------------------------------------------------------- OUTDIAG2
            '''
            DO_ELE_OIL_PUMP
            Pin       : 12
            MSb       : Byte [bit] = OUTDIAG2 [7]
            LSb       : Byte [bit] = OUTDIAG2 [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 6])
            present_sheet.cell(row=x + 1, column=36).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=36).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=36).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=36).fill = RED_FILL
                present_sheet.cell(row=1, column=36).fill = RED_FILL

            '''
            DO_RCL
            Pin       : 11
            MSb       : Byte [bit] = OUTDIAG2 [5]
            LSb       : Byte [bit] = OUTDIAG2 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 4])
            present_sheet.cell(row=x + 1, column=37).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=37).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=37).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=37).fill = RED_FILL
                present_sheet.cell(row=1, column=37).fill = RED_FILL

            '''
            DO_EVVTPWM
            Pin       : 10
            MSb       : Byte [bit] = OUTDIAG2 [3]
            LSb       : Byte [bit] = OUTDIAG2 [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 2])
            present_sheet.cell(row=x + 1, column=23).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=23).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=23).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=23).fill = RED_FILL
                present_sheet.cell(row=1, column=23).fill = RED_FILL

            '''
            DO_WGPWM
            Pin       : 9
            MSb       : Byte [bit] = OUTDIAG2 [1]
            LSb       : Byte [bit] = OUTDIAG2 [0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=7).value[7 - 0])
            present_sheet.cell(row=x + 1, column=21).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=21).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=21).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=21).fill = RED_FILL
                present_sheet.cell(row=1, column=21).fill = RED_FILL

            #---------------------------------------------------------- OUTDIAG1
            '''
            DO_IVVTPWM
            Pin       : 8
            MSb       : Byte [bit] = OUTDIAG1 [7]
            LSb       : Byte [bit] = OUTDIAG1 [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 6])
            present_sheet.cell(row=x + 1, column=18).value = \
            translate_tle_pushpull(msb, lsb)
            if present_sheet.cell(row=x + 1, column=18).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=18).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=18).fill = RED_FILL
                present_sheet.cell(row=1, column=18).fill = RED_FILL

            '''
            LSHPWM_UP
            Pin       : 7
            MSb       : Byte [bit] = OUTDIAG1 [5]
            LSb       : Byte [bit] = OUTDIAG1 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 4])
            present_sheet.cell(row=x + 1, column=19).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=19).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=19).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=19).fill = RED_FILL
                present_sheet.cell(row=1, column=19).fill = RED_FILL

            '''
            LSHPWM_DN
            Pin       : 6
            MSb       : Byte [bit] = OUTDIAG1 [3]
            LSb       : Byte [bit] = OUTDIAG1 [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 2])
            present_sheet.cell(row=x + 1, column=20).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=20).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=20).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=20).fill = RED_FILL
                present_sheet.cell(row=1, column=20).fill = RED_FILL

            '''
            CPPWM
            Pin       : 5
            MSb       : Byte [bit] = 1[1]
            LSb       : Byte [bit] = 1[0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=8).value[7 - 0])
            present_sheet.cell(row=x + 1, column=22).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=22).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=22).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=22).fill = RED_FILL
                present_sheet.cell(row=1, column=22).fill = RED_FILL

            #---------------------------------------------------------- OUTDIAG0
            '''
            IV1_CYL1
            Pin       : 1
            MSb       : Byte [bit] = OUTDIAG0 [1]
            LSb       : Byte [bit] = OUTDIAG0 [0]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 1])
            lsb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 0])
            present_sheet.cell(row=x + 1, column=14).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=14).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=14).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=14).fill = RED_FILL
                present_sheet.cell(row=1, column=14).fill = RED_FILL

            '''
            IV2_CYL3
            Pin       : 2
            MSb       : Byte [bit] = OUTDIAG0 [3]
            LSb       : Byte [bit] = OUTDIAG0 [2]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 3])
            lsb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 2])
            present_sheet.cell(row=x + 1, column=15).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=15).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=15).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=15).fill = RED_FILL
                present_sheet.cell(row=1, column=15).fill = RED_FILL

            '''
            IV3_CYL4
            Pin       : 3
            MSb       : Byte [bit] = OUTDIAG0 [5]
            LSb       : Byte [bit] = OUTDIAG0 [4]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 5])
            lsb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 4])
            present_sheet.cell(row=x + 1, column=16).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=16).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=16).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=16).fill = RED_FILL
                present_sheet.cell(row=1, column=16).fill = RED_FILL

            '''
            IV4_CYL2
            Pin       : 4
            MSb       : Byte [bit] = OUTDIAG0 [7]
            LSb       : Byte [bit] = OUTDIAG0 [6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 7])
            lsb = int(present_sheet.cell(row=x + 1, column=9).value[7 - 6])
            present_sheet.cell(row=x + 1, column=17).value = \
            translate_tle_lowside(msb, lsb)
            if present_sheet.cell(row=x + 1, column=17).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=17).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=17).fill = RED_FILL
                present_sheet.cell(row=1, column=17).fill = RED_FILL

        #----------------------------------------------------------------- Diag0
        '''
        Diag0_OCT
        OCT       : Byte[bit] = Diag0[0]
        '''
        for x in range(1, EXCEL_VERT_SCOPE - 8):

            if main_sheet.cell(row=x + 9, column=194).value != None \
            and len(str(main_sheet.cell(row=x + 9, column=194).value)) == 8 \
            and int(str(main_sheet.cell(row=x + 9,
                                        column=194).value)[7 - 0]) == 0:

                present_sheet.cell(row=x + 1, column=38).value = 'No Over Temp.'
                present_sheet.cell(row=x + 1, column=38).fill = GREEN_FILL

            elif main_sheet.cell(row=x + 9, column=194).value != None \
            and len(str(main_sheet.cell(row=x + 9, column=194).value)) == 8 \
            and int(str(main_sheet.cell(row=x + 9,
                                        column=194).value)[7 - 0]) == 1:

                present_sheet.cell(row=x + 1, column=38).value = 'Over Temp.'
                present_sheet.cell(row=x + 1, column=38).fill = RED_FILL
                present_sheet.cell(row=1, column=38).fill = RED_FILL

            else:
                present_sheet.cell(row=x + 1, column=38).value = 'No Diag0'
                present_sheet.cell(row=x + 1, column=38).fill = YELLOW_FILL

        #----------------------------------------------- Coloring and Statistics
        present_sheet.freeze_panes = 'B2'

        centr_align(present_sheet, 'A1:AM{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:AM1')

        for column in Excel_std_column_names[:9]:
            present_sheet.column_dimensions[column].width = 16
        for column in Excel_std_column_names[9:39]:
            present_sheet.column_dimensions[column].width = 20

        #=======================================================================
        # ETC Registries
        #=======================================================================
        #------------------------------------------------------------ Processing
        # print("ETC Registries")
        present_sheet = workbook['ETC Registers']

        present_sheet.cell(row=1, column=2).value = 'OVERCURRENT'
        present_sheet.cell(row=1, column=3).value = 'STATE RESP. 1'
        present_sheet.cell(row=1, column=4).value = 'STATE RESP. 2'
        present_sheet.cell(row=1, column=5).value = 'STATE RESP. 3'
        present_sheet.cell(row=1, column=6).value = 'VDD_OV_UV'

        present_sheet.cell(row=1, column=7).value = 'OUT1_H'
        present_sheet.cell(row=1, column=8).value = 'OUT0_H'
        present_sheet.cell(row=1, column=9).value = 'OUT1_L'
        present_sheet.cell(row=1, column=10).value = 'OUT0_L'

        present_sheet.cell(row=1, column=11).value = 'NDIS'
        present_sheet.cell(row=1, column=12).value = 'DIS'
        present_sheet.cell(row=1, column=13).value = 'BRIDGE'
        present_sheet.cell(row=1, column=14).value = 'HWSC_LBIST'

        present_sheet.cell(row=1, column=15).value = 'VPS_UV_REG'
        present_sheet.cell(row=1, column=16).value = 'NGFAIL'
        present_sheet.cell(row=1, column=17).value = 'ILIM_REG'
        present_sheet.cell(row=1, column=18).value = 'VDD_OV_REG'

        present_sheet.cell(row=1, column=19).value = 'VDD_UV_REG'
        present_sheet.cell(row=1, column=20).value = 'VPS_UV_REG'
        present_sheet.cell(row=1, column=21).value = 'OTSDcnt'
        present_sheet.cell(row=1, column=22).value = 'OT_WARN'

        present_sheet.cell(row=1, column=23).value = 'OT_WARN_REG'
        present_sheet.cell(row=1, column=24).value = 'NOTSD'
        present_sheet.cell(row=1, column=25).value = 'NOTSD_REG'

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=2).value = \
            main_sheet.cell(row=x + 9, column=136).value

            present_sheet.cell(row=x + 1, column=3).value = \
            main_sheet.cell(row=x + 9, column=137).value

            present_sheet.cell(row=x + 1, column=4).value = \
            main_sheet.cell(row=x + 9, column=138).value

            present_sheet.cell(row=x + 1, column=5).value = \
            main_sheet.cell(row=x + 9, column=139).value

            present_sheet.cell(row=x + 1, column=6).value = \
            main_sheet.cell(row=x + 9, column=140).value

        #-----------------------------------------------------------------------
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            '''
            OUT1_H
            MSb       : OVERCURRENT[5]
            LSb       : OVERCURRENT[6]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[5])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[6])
            present_sheet.cell(row=x + 1, column=7).value = \
            translate_etc(msb, lsb)
            if present_sheet.cell(row=x + 1, column=7).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=7).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=7).fill = RED_FILL
                present_sheet.cell(row=1, column=7).fill = RED_FILL

            '''
            OUT0_H
            MSb       : OVERCURRENT[8]
            LSb       : OVERCURRENT[9]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[8])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[9])
            present_sheet.cell(row=x + 1, column=8).value = \
            translate_etc(msb, lsb)
            if present_sheet.cell(row=x + 1, column=8).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=8).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=8).fill = RED_FILL
                present_sheet.cell(row=1, column=8).fill = RED_FILL

            '''
            OUT1_L
            MSb       : OVERCURRENT[11]
            LSb       : OVERCURRENT[12]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[11])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[12])
            present_sheet.cell(row=x + 1, column=9).value = \
            translate_etc(msb, lsb)
            if present_sheet.cell(row=x + 1, column=9).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=9).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=9).fill = RED_FILL
                present_sheet.cell(row=1, column=9).fill = RED_FILL

            '''
            OUT0_L
            MSb       : OVERCURRENT[14]
            LSb       : OVERCURRENT[15]
            '''
            msb = int(present_sheet.cell(row=x + 1, column=2).value[14])
            lsb = int(present_sheet.cell(row=x + 1, column=2).value[15])
            present_sheet.cell(row=x + 1, column=10).value = \
            translate_etc(msb, lsb)
            if present_sheet.cell(row=x + 1, column=10).value == 'No Problem':
                present_sheet.cell(row=x + 1, column=10).fill = GREEN_FILL
            else:
                present_sheet.cell(row=x + 1, column=10).fill = RED_FILL
                present_sheet.cell(row=1, column=10).fill = RED_FILL

            '''
            NDIS
            Related to: STATE RESP. 1[4]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[4] == 1:
                present_sheet.cell(row=x + 1, column=11).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=11).value = 'Low'
            present_sheet.cell(row=x + 1, column=11).fill = YELLOW_FILL

            '''
            DIS
            Related to: STATE RESP. 1[5]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[5] == 1:
                present_sheet.cell(row=x + 1, column=12).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=12).value = 'Low'
            present_sheet.cell(row=x + 1, column=12).fill = YELLOW_FILL

            '''
            BRIDGE
            Related to: STATE RESP. 1[6]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[6] == 1:
                present_sheet.cell(row=x + 1, column=13).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=13).value = 'Low'
            present_sheet.cell(row=x + 1, column=13).fill = YELLOW_FILL

            '''
            HWSC_LBIST
            Related to: STATE RESP. 1[7][8][9]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[7] == 0:
                present_sheet.cell(row=x + 1, column=14).value = 'Undone!'
            elif present_sheet.cell(row=x + 1, column=3).value[7] == 1 and \
                 present_sheet.cell(row=x + 1, column=3).value[8] == 0 and \
                 present_sheet.cell(row=x + 1, column=3).value[9] == 0:
                present_sheet.cell(row=x + 1, column=14).value = 'Fail Fail'
            elif present_sheet.cell(row=x + 1, column=3).value[7] == 1 and \
                 present_sheet.cell(row=x + 1, column=3).value[8] == 0 and \
                 present_sheet.cell(row=x + 1, column=3).value[9] == 1:
                present_sheet.cell(row=x + 1, column=14).value = 'Running Pass'
            elif present_sheet.cell(row=x + 1, column=3).value[7] == 1 and \
                 present_sheet.cell(row=x + 1, column=3).value[8] == 1 and \
                 present_sheet.cell(row=x + 1, column=3).value[9] == 0:
                present_sheet.cell(row=x + 1, column=14).value = 'Fail Pass'
            else:
                present_sheet.cell(row=x + 1, column=14).value = 'Pass Pass'
            if present_sheet.cell(row=x + 1, column=14).value == 'Pass Pass':
                present_sheet.cell(row=x + 1, column=14).fill = GREEN_FILL
            elif present_sheet.cell(row=x + 1,
                                    column=14).value == 'Running Pass':
                present_sheet.cell(row=x + 1, column=14).fill = GREEN_FILL
            elif present_sheet.cell(row=x + 1, column=14).value == 'Undone':
                present_sheet.cell(row=x + 1, column=14).fill = YELLOW_FILL
            else:
                present_sheet.cell(row=x + 1, column=14).fill = RED_FILL
                present_sheet.cell(row=1, column=14).fill = RED_FILL

            '''
            VPS_UV_REG
            Related to: STATE RESP. 1[10]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[10] == 1:
                present_sheet.cell(row=x + 1, column=15).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=15).value = 'Low'
            present_sheet.cell(row=x + 1, column=15).fill = YELLOW_FILL

            '''
            NGFAIL
            Related to: STATE RESP. 1[11]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[11] == 1:
                present_sheet.cell(row=x + 1, column=16).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=16).value = 'Low'
            present_sheet.cell(row=x + 1, column=16).fill = YELLOW_FILL

            '''
            ILIM_REG
            Related to: STATE RESP. 1[12]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[12] == 1:
                present_sheet.cell(row=x + 1, column=17).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=17).value = 'Low'
            present_sheet.cell(row=x + 1, column=17).fill = YELLOW_FILL

            '''
            VDD_OV_REG
            Related to: STATE RESP. 1[13]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[13] == 1:
                present_sheet.cell(row=x + 1, column=18).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=18).value = 'Low'
            present_sheet.cell(row=x + 1, column=18).fill = YELLOW_FILL

            '''
            VDD_UV_REG
            Related to: STATE RESP. 1[14]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[14] == 1:
                present_sheet.cell(row=x + 1, column=19).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=19).value = 'Low'
            present_sheet.cell(row=x + 1, column=19).fill = YELLOW_FILL

            '''
            VPS_UV_REG
            Related to: STATE RESP. 1[15]
            '''
            if present_sheet.cell(row=x + 1, column=3).value[15] == 1:
                present_sheet.cell(row=x + 1, column=20).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=20).value = 'Low'
            present_sheet.cell(row=x + 1, column=20).fill = YELLOW_FILL

            '''
            OTSDcnt
            Related to: STATE RESP. 2[4][5][6][7][8][9]
            '''
            present_sheet.cell(row=x + 1, column=21).value = \
            int(present_sheet.cell(row=x + 1, column=4).value[4]) * 32 + \
            int(present_sheet.cell(row=x + 1, column=4).value[5]) * 16 + \
            int(present_sheet.cell(row=x + 1, column=4).value[6]) * 8 + \
            int(present_sheet.cell(row=x + 1, column=4).value[7]) * 4 + \
            int(present_sheet.cell(row=x + 1, column=4).value[8]) * 2 + \
            int(present_sheet.cell(row=x + 1, column=4).value[9])
            present_sheet.cell(row=x + 1, column=21).fill = YELLOW_FILL

            '''
            OT_WARN
            Related to: STATE RESP. 2[10]
            '''
            if present_sheet.cell(row=x + 1, column=4).value[10] == 1:
                present_sheet.cell(row=x + 1, column=22).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=22).value = 'Low'
            present_sheet.cell(row=x + 1, column=22).fill = YELLOW_FILL

            '''
            OT_WARN_REG
            Related to: STATE RESP. 1[11]
            '''
            if present_sheet.cell(row=x + 1, column=4).value[11] == 1:
                present_sheet.cell(row=x + 1, column=23).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=23).value = 'Low'
            present_sheet.cell(row=x + 1, column=23).fill = YELLOW_FILL

            '''
            NOTSD
            Related to: STATE RESP. 1[12]
            '''
            if present_sheet.cell(row=x + 1, column=4).value[12] == 1:
                present_sheet.cell(row=x + 1, column=24).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=24).value = 'Low'
            present_sheet.cell(row=x + 1, column=24).fill = YELLOW_FILL

            '''
            NOTSD_REG
            Related to: STATE RESP. 1[13]
            '''
            if present_sheet.cell(row=x + 1, column=4).value[13] == 1:
                present_sheet.cell(row=x + 1, column=25).value = 'High'
            else:
                present_sheet.cell(row=x + 1, column=25).value = 'Low'
            present_sheet.cell(row=x + 1, column=25).fill = YELLOW_FILL

        #----------------------------------------------- Coloring and Statistics
        present_sheet.freeze_panes = 'B2'

        centr_align(present_sheet, 'A1:Z{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:Z1')

        for column in Excel_std_column_names[:1]:
            present_sheet.column_dimensions[column].width = 16

        for column in Excel_std_column_names[1:6]:
            present_sheet.column_dimensions[column].width = 20

        for column in Excel_std_column_names[6:26]:
            present_sheet.column_dimensions[column].width = 16

        for cell in present_sheet.iter_rows(min_row=1,
                                            max_row=1,
                                            min_col=6,
                                            max_col=26):
            for j in range(len(cell)):
                if j % 2 == 0:
                    cell[j].fill = LGREY_FILL
                else:
                    cell[j].fill = GREY_FILL

        #=======================================================================
        # KP254 Values
        #=======================================================================
        #------------------------------------------------------------ Processing
        # print("KP254 Values")
        present_sheet = workbook['KP254 Values']

        present_sheet.cell(row=1, column=12).value = 'KP254_PRESSURE [kPa]'
        present_sheet.cell(row=1, column=13).value = 'KP254_TEMP [°C]'
        # present_sheet.cell(row=1, column=14).value = 'KP254_DIAG_H'
        # present_sheet.cell(row=1, column=15).value = 'KP254_DIAG_L'
        present_sheet.cell(row=1, column=14).value = 'Diagnosis'

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=11).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=11).fill = BROWN_FILL
                present_sheet.cell(row=1, column=11).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=11).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=11).fill = YELLOW_FILL

        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=12).value = \
            byte_to_pressure(main_sheet.cell(row=x + 9, column=144).value)

            present_sheet.cell(row=x + 1, column=13).value = \
            byte_to_temperature(main_sheet.cell(row=x + 9, column=145).value)

            # present_sheet.cell(row=x + 1, column=14).value = \
            # main_sheet.cell(row=x + 9, column=146).value
            #
            # present_sheet.cell(row=x + 1, column=15).value = \
            # main_sheet.cell(row=x + 9, column=147).value

            KP254_diagnosis(present_sheet,
                            x + 1,
                            14,
                            main_sheet.cell(row=x + 9, column=146).value)

        #----------------------------------------------------------------- Chart
        chart = LineChart()

        data = Reference(worksheet=present_sheet,
                           min_row=1,
                           max_row=EXCEL_VERT_SCOPE,
                           min_col=12,
                           max_col=13)
        chart.add_data(data, titles_from_data=True)

        cats = Reference(worksheet=present_sheet,
                           min_row=2,
                           max_row=EXCEL_VERT_SCOPE,
                           min_col=11,
                           max_col=11)
        chart.set_categories(cats)

        chart.x_axis.title = "Sample Time"
        chart.y_axis.title = "KP254 Values"

        chart.width = CHART_WIDTH
        chart.height = CHART_HIGHT

        present_sheet.add_chart(chart, "A5")

        #----------------------------------------------- Coloring and Statistics
        centr_align(present_sheet, 'A1:O{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:O1')

        for column in Excel_std_column_names[:16]:
            present_sheet.column_dimensions[column].width = 18

        present_sheet.column_dimensions[Excel_std_column_names[9]].width = 4

        #=======================================================================
        # Temperature Values
        #=======================================================================
        #------------------------------------------------------------ Processing
        # print("Temperature Values")
        present_sheet = workbook['Temperature Values']

        present_sheet.cell(row=1, column=11).value = 'MOSFET [°C]'
        present_sheet.cell(row=1, column=12).value = 'SBC (TLE) [°C]'
        present_sheet.cell(row=1, column=13).value = 'Chamber [°C]'
        present_sheet.cell(row=1, column=14).value = 'Water Pump [°C]'
        present_sheet.cell(row=1, column=15).value = 'MCU [°C]'

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=10).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=10).fill = BROWN_FILL
                present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=10).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=10).fill = YELLOW_FILL

        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=11).value = \
            main_sheet.cell(row=x + 9, column=148).value

            present_sheet.cell(row=x + 1, column=12).value = \
            main_sheet.cell(row=x + 9, column=149).value

            present_sheet.cell(row=x + 1, column=13).value = \
            main_sheet.cell(row=x + 9, column=150).value

            present_sheet.cell(row=x + 1, column=14).value = \
            main_sheet.cell(row=x + 9, column=151).value

            present_sheet.cell(row=x + 1, column=15).value = \
            main_sheet.cell(row=x + 9, column=157).value

        #----------------------------------------------------------------- Chart
        chart = LineChart()

        data = Reference(worksheet=present_sheet,
                                  min_row=1,
                                  max_row=EXCEL_VERT_SCOPE,
                                  min_col=11,
                                  max_col=15)
        chart.add_data(data, titles_from_data=True)

        cats = Reference(worksheet=present_sheet,
                          min_row=2,
                          max_row=EXCEL_VERT_SCOPE,
                          min_col=10,
                          max_col=10)
        chart.set_categories(cats)

        chart.x_axis.title = "Sample Time"
        chart.y_axis.title = "Temperature Values"

        chart.width = CHART_WIDTH
        chart.height = CHART_HIGHT

        present_sheet.add_chart(chart, "A5")

        #----------------------------------------------- Coloring and Statistics
        centr_align(present_sheet, 'A1:O{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:O1')
        for column in Excel_std_column_names[1:14]:
            present_sheet.column_dimensions[column].width = 18

        #=======================================================================
        # Reset Values
        #=======================================================================
        '''
        [Internal Status Report Response]
        http://192.168.5.62:8090/display/ECMT2/
        0x00+-+Internal+Status+Report+Response
        '''

        #------------------------------------------------------------ Processing
        present_sheet = workbook['Reset Registers']
        # print("Reset Values")

        present_sheet.cell(row=1, column=2).value = 'Mode'
        present_sheet.cell(row=1, column=3).value = 'MC_LastRST'
        present_sheet.cell(row=1, column=4).value = 'MU_LastRST'
        present_sheet.cell(row=1, column=5).value = 'MU_RstOutMCCont'
        present_sheet.cell(row=1, column=6).value = 'MU_ErrCode'
        present_sheet.cell(row=1, column=7).value = 'MU_State'
        present_sheet.cell(row=1, column=8).value = 'MC_ExRstCont'

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=2).value = \
            main_sheet.cell(row=x + 9, column=195).value

            if int(present_sheet.cell(row=x + 1, column=2).value) == 0:
                present_sheet.cell(row=x + 1, column=2).value = 'Stand-By Mode'

            elif int(present_sheet.cell(row=x + 1, column=2).value) == 1:
                present_sheet.cell(row=x + 1, column=2).value = 'Idle Mode'

            elif int(present_sheet.cell(row=x + 1, column=2).value) == 2:
                present_sheet.cell(row=x + 1, column=2).value = 'Part-load Mode'

            elif int(present_sheet.cell(row=x + 1, column=2).value) == 3:
                present_sheet.cell(row=x + 1, column=2).value = 'Full-load Mode'

            elif int(present_sheet.cell(row=x + 1, column=2).value) == 4:
                present_sheet.cell(row=x + 1, column=2).value = 'The Worst Case'

            else:
                present_sheet.cell(row=x + 1, column=2).value = 'Invalid!'
                present_sheet.cell(row=x + 1, column=2).fill = YELLOW_FILL
                present_sheet.cell(row=1, column=2).fill = YELLOW_FILL

            present_sheet.cell(row=x + 1, column=3).value = \
            main_sheet.cell(row=x + 9, column=152).value

            if int(present_sheet.cell(row=x + 1, column=3).value) == 128:
                present_sheet.cell(row=x + 1, column=3).value = 'PowerOnRst'

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 64:
                present_sheet.cell(row=x + 1, column=3).value = 'ExtRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 32:
                present_sheet.cell(row=x + 1, column=3).value = 'LossOfLockRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 16:
                present_sheet.cell(row=x + 1, column=3).value = 'LossOfClockRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 8:
                present_sheet.cell(row=x + 1, column=3).value = 'WatchdogRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 4:
                present_sheet.cell(row=x + 1, column=3).value = 'CheckStopRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 2:
                present_sheet.cell(row=x + 1, column=3).value = 'SW_WatchdogRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=3).value) == 1:
                present_sheet.cell(row=x + 1, column=3).value = 'SW_SystemRst'
                present_sheet.cell(row=x + 1, column=3).fill = RED_FILL
                present_sheet.cell(row=1, column=3).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            else:
                present_sheet.cell(row=x + 1, column=3).value = 'Invalid!'
                present_sheet.cell(row=x + 1, column=3).fill = YELLOW_FILL
                present_sheet.cell(row=1, column=3).fill = YELLOW_FILL

            present_sheet.cell(row=x + 1, column=4).value = \
            main_sheet.cell(row=x + 9, column=153).value

            if int(present_sheet.cell(row=x + 1, column=4).value) == 1:
                present_sheet.cell(row=x + 1, column=4).value = 'PowerOnRst'

            elif int(present_sheet.cell(row=x + 1, column=4).value) == 2:
                present_sheet.cell(row=x + 1, column=4).value = 'ExtRst'
                present_sheet.cell(row=x + 1, column=4).fill = RED_FILL
                present_sheet.cell(row=1, column=4).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=4).value) == 3:
                present_sheet.cell(row=x + 1, column=4).value = 'IntRst'
                present_sheet.cell(row=x + 1, column=4).fill = RED_FILL
                present_sheet.cell(row=1, column=4).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=4).value) == 4:
                present_sheet.cell(row=x + 1, column=4).value = 'MU_RstOut'
                present_sheet.cell(row=x + 1, column=4).fill = RED_FILL
                present_sheet.cell(row=1, column=4).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=4).value) == 5:
                present_sheet.cell(row=x + 1, column=4).value = 'MC_RstRequest'
                present_sheet.cell(row=x + 1, column=4).fill = RED_FILL
                present_sheet.cell(row=1, column=4).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            else:
                present_sheet.cell(row=x + 1, column=4).value = 'Invalid!'
                present_sheet.cell(row=x + 1, column=4).fill = YELLOW_FILL
                present_sheet.cell(row=1, column=4).fill = YELLOW_FILL

            present_sheet.cell(row=x + 1, column=5).value = \
            main_sheet.cell(row=x + 9, column=154).value

            if int(main_sheet.cell(row=x + 9, column=154).value) > 0:
                present_sheet.cell(row=x + 1, column=5).fill = RED_FILL
                present_sheet.cell(row=1, column=5).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            present_sheet.cell(row=x + 1, column=6).value = \
            int(main_sheet.cell(row=x + 9, column=155).value)

            if int(present_sheet.cell(row=x + 1, column=6).value) == 0:
                present_sheet.cell(row=x + 1, column=6).value = 'NO_ERR'

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 1:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_BIST'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 2:
                present_sheet.cell(row=x + 1, column=6).value = 'OVD_UVD'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 3:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_ADDR'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 4:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_COMP'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 5:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_FS_IST'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 6:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_RED_SOP'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 7:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_PFM_CPL'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 8:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_PFM_0'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 9:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_PFM_1'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 10:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_PFM_2'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 11:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_PFM_3'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 12:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_IGK_CPL'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 13:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_EOLT'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=6).value) == 14:
                present_sheet.cell(row=x + 1, column=6).value = 'MU_DI_RST'
                present_sheet.cell(row=x + 1, column=6).fill = RED_FILL
                present_sheet.cell(row=1, column=6).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            else:
                present_sheet.cell(row=x + 1, column=6).value = 'Invalid!'
                present_sheet.cell(row=x + 1, column=6).fill = YELLOW_FILL
                present_sheet.cell(row=1, column=6).fill = YELLOW_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = YELLOW_FILL

            present_sheet.cell(row=x + 1, column=7).value = \
            main_sheet.cell(row=x + 9, column=156).value

            if int(present_sheet.cell(row=x + 1, column=7).value) == 3:
                present_sheet.cell(row=x + 1, column=7).value = 'INIT'

            elif int(present_sheet.cell(row=x + 1, column=7).value) == 6:
                present_sheet.cell(row=x + 1, column=7).value = 'CONFIG'

            elif int(present_sheet.cell(row=x + 1, column=7).value) == 10:
                present_sheet.cell(row=x + 1, column=7).value = 'NORMAL'

            elif int(present_sheet.cell(row=x + 1, column=7).value) == 9:
                present_sheet.cell(row=x + 1, column=7).value = 'ELOT'

            elif int(present_sheet.cell(row=x + 1, column=7).value) == 14:
                present_sheet.cell(row=x + 1, column=7).value = 'DISABLE'

            else:
                present_sheet.cell(row=x + 1, column=7).value = 'DISABLE'
                present_sheet.cell(row=x + 1, column=7).fill = YELLOW_FILL

            present_sheet.cell(row=x + 1, column=8).value = \
            main_sheet.cell(row=x + 9, column=196).value

            if int(present_sheet.cell(row=x + 1, column=8).value) > \
            int(present_sheet.cell(row=2, column=8).value):
                present_sheet.cell(row=x + 1, column=8).fill = RED_FILL
                present_sheet.cell(row=1, column=8).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #----------------------------------------------- Coloring and Statistics
        centr_align(present_sheet, 'A1:I{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:I1')
        for column in Excel_std_column_names[:9]:
            present_sheet.column_dimensions[column].width = 20

        present_sheet.freeze_panes = "A2"

        #=======================================================================
        # Monitoring Unit
        #=======================================================================
        '''
        [Monitoring Unit Report Request]
        http://192.168.5.62:8090/display/ECMT2/
        0x40+-+Monitoring+Unit+Report+Request
        '''

        # print("Monitoring Unit")
        present_sheet = workbook['Monitoring Unit']

        #------------------------------------------------------------------ Time
        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

        #----------------------------------------------------------- MU Byte 1&2
        present_sheet.cell(row=1, column=3).value = 'Word 1'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=3).value = \
            main_sheet.cell(row=x + 9, column=158).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=19).value = 'CTR Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=19).value = \
            main_sheet.cell(row=x + 9, column=158).value[-12:-10]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=20).value = 'MUX Validity Check'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=20).value = \
            main_sheet.cell(row=x + 9, column=158).value[-10:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=21).value = 'PVS2 Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=21).value = \
            convert_digitl_to_analog(binary_string_to_decimal
                                  (main_sheet.cell
                                   (row=x + 9, column=158).value[-8:]), 5, 8)
            if abs(present_sheet.cell(row=x + 1, column=21).value - 0.75) > 0.08:
                present_sheet.cell(row=x + 1, column=21).fill = RED_FILL
                present_sheet.cell(row=1, column=21).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #----------------------------------------------------------- MU Byte 3&4
        present_sheet.cell(row=1, column=4).value = 'Word 2'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=4).value = \
            main_sheet.cell(row=x + 9, column=159).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=22).value = 'ABC FS-IST'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=22).value = \
            main_sheet.cell(row=x + 9, column=159).value[-9:-6]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=23).value = 'ABC Answer CPL Check'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=23).value = \
            main_sheet.cell(row=x + 9, column=159).value[-6:-3]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=24).value = 'ABC IGK CPL Check'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=24).value = \
            main_sheet.cell(row=x + 9, column=159).value[-3:]

        #----------------------------------------------------------- MU Byte 5&6
        present_sheet.cell(row=1, column=5).value = 'Word 3'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=5).value = \
            main_sheet.cell(row=x + 9, column=160).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=25).value = 'ABC PFM2'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=25).value = \
            main_sheet.cell(row=x + 9, column=160).value[-12:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=26).value = 'ABC PFM1'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=26).value = \
            main_sheet.cell(row=x + 9, column=160).value[-8:-4]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=27).value = 'ABC PFM0'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=27).value = \
            main_sheet.cell(row=x + 9, column=160).value[-4:]

        #----------------------------------------------------------- MU Byte 7&8
        present_sheet.cell(row=1, column=6).value = 'Word 4'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=6).value = \
            main_sheet.cell(row=x + 9, column=161).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=28).value = 'ABC State Transition'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=28).value = \
            main_sheet.cell(row=x + 9, column=161).value[-7:-4]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=29).value = 'ABC PFM3'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=29).value = \
            main_sheet.cell(row=x + 9, column=161).value[-4:]

        #---------------------------------------------------------- MU Byte 9&10
        present_sheet.cell(row=1, column=7).value = 'Word 5'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=7).value = \
            main_sheet.cell(row=x + 9, column=162).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=30).value = 'Timer FS-IST'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=30).value = \
            main_sheet.cell(row=x + 9, column=162).value[-12:]

        #--------------------------------------------------------- MU Byte 11&12
        present_sheet.cell(row=1, column=8).value = 'Word 6'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=8).value = \
            main_sheet.cell(row=x + 9, column=163).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=31).value = 'Timer Injection Off'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=31).value = \
            main_sheet.cell(row=x + 9, column=163).value[-10:-5]

        #--------------------------------------------------------- MU Byte 13&14
        present_sheet.cell(row=1, column=9).value = 'Word 7'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=9).value = \
            main_sheet.cell(row=x + 9, column=164).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=32).value = 'Timer PFM0'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=32).value = \
            main_sheet.cell(row=x + 9, column=164).value[-8:]

        #--------------------------------------------------------- MU Byte 15&16
        present_sheet.cell(row=1, column=10).value = 'Word 8'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=10).value = \
            main_sheet.cell(row=x + 9, column=165).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=33).value = 'Timer PFM1'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=33).value = \
            main_sheet.cell(row=x + 9, column=165).value[-9:]

        #--------------------------------------------------------- MU Byte 17&18
        present_sheet.cell(row=1, column=11).value = 'Word 9'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=11).value = \
            main_sheet.cell(row=x + 9, column=166).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=34).value = 'Timer PFM2'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=34).value = \
            main_sheet.cell(row=x + 9, column=166).value[-9:]

        #--------------------------------------------------------- MU Byte 19&20
        present_sheet.cell(row=1, column=12).value = 'Word 10'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=12).value = \
            main_sheet.cell(row=x + 9, column=167).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=35).value = 'Timer PFM3'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=35).value = \
            main_sheet.cell(row=x + 9, column=167).value[-12:]

        #--------------------------------------------------------- MU Byte 21&22
        present_sheet.cell(row=1, column=13).value = 'Word 11'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=13).value = \
            main_sheet.cell(row=x + 9, column=168).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=36).value = 'Reset Counter MU'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=36).value = \
            main_sheet.cell(row=x + 9, column=168).value[-12:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=37).value = 'Tool0'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=37).value = \
            main_sheet.cell(row=x + 9, column=168).value[-8:-7]
        if int(present_sheet.cell(row=x + 1, column=37).value == 0):
            present_sheet.cell(row=x + 1, column=37).fill = RED_FILL
            present_sheet.cell(row=1, column=37).fill = RED_FILL
            for k in range(157, 174):
                main_sheet.cell(row=7, column=k).fill = RED_FILL

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=38).value = 'PEN'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=38).value = \
            main_sheet.cell(row=x + 9, column=168).value[-7:-6]
        if int(present_sheet.cell(row=x + 1, column=38).value == 0):
            present_sheet.cell(row=x + 1, column=38).fill = RED_FILL
            present_sheet.cell(row=1, column=38).fill = RED_FILL
            for k in range(157, 174):
                main_sheet.cell(row=7, column=k).fill = RED_FILL

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=39).value = 'NDIS1'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=39).value = \
            main_sheet.cell(row=x + 9, column=168).value[-6:-5]
        if int(present_sheet.cell(row=x + 1, column=39).value == 0):
            present_sheet.cell(row=x + 1, column=39).fill = RED_FILL
            present_sheet.cell(row=1, column=39).fill = RED_FILL
            for k in range(157, 174):
                main_sheet.cell(row=7, column=k).fill = RED_FILL

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=40).value = 'NDIS0'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=40).value = \
            main_sheet.cell(row=x + 9, column=168).value[-5:-4]
        if int(present_sheet.cell(row=x + 1, column=40).value == 0):
            present_sheet.cell(row=x + 1, column=40).fill = RED_FILL
            present_sheet.cell(row=1, column=40).fill = RED_FILL
            for k in range(157, 174):
                main_sheet.cell(row=7, column=k).fill = RED_FILL

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=41).value = 'Error Code MU'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=41).value = \
            main_sheet.cell(row=x + 9, column=168).value[-4:]

            if int(present_sheet.cell(row=x + 1, column=41).value) == 0:
                present_sheet.cell(row=x + 1, column=41).value = 'NO_ERR'

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 1:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_BIST'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 2:
                present_sheet.cell(row=x + 1, column=41).value = 'OVD_UVD'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 3:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_ADDR'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 4:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_COMP'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 5:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_FS_IST'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 6:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_RED_SOP'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 7:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_PFM_CPL'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 8:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_PFM_0'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 9:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_PFM_1'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 10:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_PFM_2'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 11:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_PFM_3'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(152, 157):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 12:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_IGK_CPL'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 13:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_EOLT'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            elif int(present_sheet.cell(row=x + 1, column=41).value) == 14:
                present_sheet.cell(row=x + 1, column=41).value = 'MU_DI_RST'
                present_sheet.cell(row=1, column=41).fill = RED_FILL
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

            else:
                present_sheet.cell(row=x + 1, column=41).value = 'Invalid!'
                for k in range(19, 54):
                    present_sheet.cell(row=1, column=k).fill = YELLOW_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = YELLOW_FILL

        #--------------------------------------------------------- MU Byte 23&24
        present_sheet.cell(row=1, column=14).value = 'Word 12'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=14).value = \
            main_sheet.cell(row=x + 9, column=169).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=42).value = 'CTR Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=42).value = \
            main_sheet.cell(row=x + 9, column=169).value[-12:-10]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=43).value = 'MUX Validity Check'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=43).value = \
            main_sheet.cell(row=x + 9, column=169).value[-10:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=44).value = 'Teack1 Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=44).value = \
            convert_digitl_to_analog(binary_string_to_decimal
                                  (main_sheet.cell
                                   (row=x + 9, column=169).value[-8:]), 5, 8)
            if abs(present_sheet.cell(row=x + 1, column=44).value - 2.5) > 0.250:
                present_sheet.cell(row=x + 1, column=44).fill = RED_FILL
                present_sheet.cell(row=1, column=44).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #--------------------------------------------------------- MU Byte 25&26
        present_sheet.cell(row=1, column=15).value = 'Word 13'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=15).value = \
            main_sheet.cell(row=x + 9, column=170).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=45).value = 'CTR Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=45).value = \
            main_sheet.cell(row=x + 9, column=170).value[-12:-10]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=46).value = 'MUX Validity Check'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=46).value = \
            main_sheet.cell(row=x + 9, column=170).value[-10:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=47).value = 'Teack2 Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=47).value = \
            convert_digitl_to_analog(binary_string_to_decimal
                                  (main_sheet.cell
                                   (row=x + 9, column=170).value[-8:]), 5, 8)
            if abs(present_sheet.cell(row=x + 1, column=47).value - 2.5) > 0.250:
                present_sheet.cell(row=x + 1, column=47).fill = RED_FILL
                present_sheet.cell(row=1, column=47).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #--------------------------------------------------------- MU Byte 27&28
        present_sheet.cell(row=1, column=16).value = 'Word 14'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=16).value = \
            main_sheet.cell(row=x + 9, column=171).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=48).value = '+5V Value'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=48).value = \
            convert_digitl_to_analog(binary_string_to_decimal
                                  (main_sheet.cell
                                   (row=x + 9, column=171).value[-8:]), 5, 8)
            if abs(present_sheet.cell(row=x + 1, column=48).value - 2.5) > 0.250:
                present_sheet.cell(row=x + 1, column=48).fill = RED_FILL
                present_sheet.cell(row=1, column=48).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #--------------------------------------------------------- MU Byte 29&30
        present_sheet.cell(row=1, column=17).value = 'Word 15'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=17).value = \
            main_sheet.cell(row=x + 9, column=172).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=49).value = 'Reset Type'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=49).value = \
            main_sheet.cell(row=x + 9, column=172).value[-12:-9]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=50).value = 'RAM March POR'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=50).value = \
            main_sheet.cell(row=x + 9, column=172).value[-9:-8]
            if int(present_sheet.cell(row=x + 1, column=50).value) != 0:
                present_sheet.cell(row=x + 1, column=50).fill = RED_FILL
                present_sheet.cell(row=1, column=50).fill = RED_FILL
                for k in range(157, 174):
                    main_sheet.cell(row=7, column=k).fill = RED_FILL

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=51).value = 'Last BIST Record POR'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=51).value = \
            main_sheet.cell(row=x + 9, column=172).value[-8:]

        #--------------------------------------------------------- MU Byte 31&32
        present_sheet.cell(row=1, column=18).value = 'Word 16'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=18).value = \
            main_sheet.cell(row=x + 9, column=173).value

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=52).value = 'Customer Release Number'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=52).value = \
            main_sheet.cell(row=x + 9, column=173).value[-12:-8]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=53).value = 'PTHW Release Number'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=53).value = \
            main_sheet.cell(row=x + 9, column=173).value[-8:-4]

        #-------------------------------------------------- MU Assessment Column
        present_sheet.cell(row=1, column=54).value = 'PTS Release Number'
        for x in range(1, EXCEL_VERT_SCOPE - 8):
            present_sheet.cell(row=x + 1, column=54).value = \
            main_sheet.cell(row=x + 9, column=173).value[-4:]

        #---------------------------------------------------- MU Entering Values
        # for x in range(1, EXCEL_VERT_SCOPE - 8):
        #     present_sheet.cell(row=x + 1, column=3).value = \
        #     main_sheet.cell(row=x + 9, column=158).value
        #
        #     present_sheet.cell(row=x + 1, column=4).value = \
        #     main_sheet.cell(row=x + 9, column=159).value
        #
        #     present_sheet.cell(row=x + 1, column=5).value = \
        #     main_sheet.cell(row=x + 9, column=160).value
        #
        #     present_sheet.cell(row=x + 1, column=6).value = \
        #     main_sheet.cell(row=x + 9, column=161).value
        #
        #     present_sheet.cell(row=x + 1, column=7).value = \
        #     main_sheet.cell(row=x + 9, column=162).value
        #
        #     present_sheet.cell(row=x + 1, column=8).value = \
        #     main_sheet.cell(row=x + 9, column=163).value
        #
        #     present_sheet.cell(row=x + 1, column=9).value = \
        #     main_sheet.cell(row=x + 9, column=164).value
        #
        #     present_sheet.cell(row=x + 1, column=10).value = \
        #     main_sheet.cell(row=x + 9, column=165).value
        #
        #     present_sheet.cell(row=x + 1, column=11).value = \
        #     main_sheet.cell(row=x + 9, column=166).value
        #
        #     present_sheet.cell(row=x + 1, column=12).value = \
        #     main_sheet.cell(row=x + 9, column=167).value
        #
        #     present_sheet.cell(row=x + 1, column=13).value = \
        #     main_sheet.cell(row=x + 9, column=168).value
        #
        #     present_sheet.cell(row=x + 1, column=14).value = \
        #     main_sheet.cell(row=x + 9, column=169).value
        #
        #     present_sheet.cell(row=x + 1, column=15).value = \
        #     main_sheet.cell(row=x + 9, column=170).value
        #
        #     present_sheet.cell(row=x + 1, column=16).value = \
        #     main_sheet.cell(row=x + 9, column=171).value
        #
        #     present_sheet.cell(row=x + 1, column=17).value = \
        #     main_sheet.cell(row=x + 9, column=172).value
        #
        #     present_sheet.cell(row=x + 1, column=18).value = \
        #     main_sheet.cell(row=x + 9, column=173).value

        #-------------------------------------------------------------- Coloring
        present_sheet.freeze_panes = 'B2'

        centr_align(present_sheet, 'A1:BB{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:BB1')

        for column in Excel_std_column_names[:1]:
            present_sheet.column_dimensions[column].width = 16

        for column in Excel_std_column_names[1:2]:
            present_sheet.column_dimensions[column].width = 12

        for column in Excel_std_column_names[2:18]:
            present_sheet.column_dimensions[column].width = 18

        for column in Excel_std_column_names[18:55]:
            present_sheet.column_dimensions[column].width = 12

        for cell in present_sheet.iter_rows(min_row=1,
                                            max_row=EXCEL_VERT_SCOPE - 8,
                                            min_col=2,
                                            max_col=18):
            for j in range(len(cell)):
                if j % 2 == 0:
                    cell[j].fill = LGREY_FILL
                else:
                    cell[j].fill = GREY_FILL

        #=======================================================================
        # Frequency Outputs
        #=======================================================================
        # print("Frequency Outputs")

        '''
        [Test Functional Modes]
        http://192.168.5.62:8090/display/ECMT2/Test+Functional+Modes
        '''

        present_sheet = workbook['Frequency Outputs']

        #------------------------------------------------------------ Processing
        present_sheet.cell(row=1, column=3).value = 'IGC1'
        present_sheet.cell(row=1, column=4).value = 'IGC2'
        present_sheet.cell(row=1, column=5).value = 'IGC3'
        present_sheet.cell(row=1, column=6).value = 'IGC4'

        present_sheet.cell(row=1, column=7).value = 'INJ1'
        present_sheet.cell(row=1, column=8).value = 'INJ2'
        present_sheet.cell(row=1, column=9).value = 'INJ3'
        present_sheet.cell(row=1, column=10).value = 'INJ4'

        present_sheet.cell(row=1, column=11).value = 'IVVT'
        present_sheet.cell(row=1, column=17).value = 'EVVT'

        present_sheet.cell(row=1, column=12).value = 'LSHDOWN'
        present_sheet.cell(row=1, column=13).value = 'LSHUP'

        present_sheet.cell(row=1, column=14).value = 'ALT_CMD'

        present_sheet.cell(row=1, column=15).value = 'WG'
        present_sheet.cell(row=1, column=16).value = 'CP'

        #=======================================================================
        # Standby Mode
        #=======================================================================

        if main_sheet.cell(row=2, column=2).value == 'standby':
            #-------------------------------------------------------------- IGCx
            # Frequency
            present_sheet.cell(row=3, column=3).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=3).value = 0

            # Frequency
            present_sheet.cell(row=3, column=4).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=4).value = 0

            # Frequency
            present_sheet.cell(row=3, column=5).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=5).value = 0

            # Frequency
            present_sheet.cell(row=3, column=6).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=6).value = 0

            #-------------------------------------------------------------- INJx
            # Frequency
            present_sheet.cell(row=3, column=7).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=7).value = 0

            # Frequency
            present_sheet.cell(row=3, column=8).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=8).value = 0

            # Frequency
            present_sheet.cell(row=3, column=9).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=9).value = 0

            # Frequency
            present_sheet.cell(row=3, column=10).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=10).value = 0

            #------------------------------------------------------- IVVT & EVVT
            # Frequency
            present_sheet.cell(row=3, column=11).value = 0
            present_sheet.cell(row=4, column=11).value = 0

            # Frequency
            present_sheet.cell(row=3, column=17).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=17).value = 0

            #-------------------------------------------------------- LSHUP/DOWN
            # Frequency
            present_sheet.cell(row=3, column=12).value = 0
            present_sheet.cell(row=4, column=12).value = 0

            # Frequency
            present_sheet.cell(row=3, column=13).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=13).value = 0

            #----------------------------------------------------------- ALT_CMD
            # Frequency
            present_sheet.cell(row=3, column=14).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=14).value = 0

            #---------------------------------------------------------------- WG
            # Frequency
            present_sheet.cell(row=3, column=15).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=15).value = 0

            #---------------------------------------------------------------- CP
            # Frequency
            present_sheet.cell(row=3, column=16).value = 0
            # Duty cycle
            present_sheet.cell(row=4, column=16).value = 0

        #=======================================================================
        # Idle Mode
        #=======================================================================

        if main_sheet.cell(row=2, column=2).value == 'idle':
            #-------------------------------------------------------------- IGCx
            # Frequency
            present_sheet.cell(row=3, column=3).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=3).value = 100 - 98.62

            # Frequency
            present_sheet.cell(row=3, column=4).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=4).value = 100 - 98.62

            # Frequency
            present_sheet.cell(row=3, column=5).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=5).value = 100 - 98.62

            # Frequency
            present_sheet.cell(row=3, column=6).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=6).value = 100 - 98.62

            #-------------------------------------------------------------- INJx
            # Frequency
            present_sheet.cell(row=3, column=7).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=7).value = 2

            # Frequency
            present_sheet.cell(row=3, column=8).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=8).value = 2

            # Frequency
            present_sheet.cell(row=3, column=9).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=9).value = 2

            # Frequency
            present_sheet.cell(row=3, column=10).value = 6
            # Duty cycle
            present_sheet.cell(row=4, column=10).value = 2

            #------------------------------------------------------- IVVT & EVVT
            # Frequency
            present_sheet.cell(row=3, column=11).value = 250
            present_sheet.cell(row=4, column=11).value = 10

            # Frequency
            present_sheet.cell(row=3, column=17).value = 250
            # Duty cycle
            present_sheet.cell(row=4, column=17).value = 10

            #-------------------------------------------------------- LSHUP/DOWN
            # Frequency
            present_sheet.cell(row=3, column=12).value = 10
            present_sheet.cell(row=4, column=12).value = 10

            # Frequency
            present_sheet.cell(row=3, column=13).value = 10
            # Duty cycle
            present_sheet.cell(row=4, column=13).value = 10

            #----------------------------------------------------------- ALT_CMD
            # Frequency
            present_sheet.cell(row=3, column=14).value = 100
            # Duty cycle
            present_sheet.cell(row=4, column=14).value = 10

            #---------------------------------------------------------------- WG
            # Frequency
            present_sheet.cell(row=3, column=15).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=15).value = 10

            #---------------------------------------------------------------- CP
            # Frequency
            present_sheet.cell(row=3, column=16).value = 15
            # Duty cycle
            present_sheet.cell(row=4, column=16).value = 20

        #=======================================================================
        # Part Load Mode
        #=======================================================================

        if main_sheet.cell(row=2, column=2).value == 'part_load':
            #-------------------------------------------------------------- IGCx
            # Frequency
            present_sheet.cell(row=3, column=3).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=3).value = 100 - 94.25

            # Frequency
            present_sheet.cell(row=3, column=4).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=4).value = 100 - 94.25

            # Frequency
            present_sheet.cell(row=3, column=5).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=5).value = 100 - 94.25

            # Frequency
            present_sheet.cell(row=3, column=6).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=6).value = 100 - 94.25

            #-------------------------------------------------------------- INJx
            # Frequency
            present_sheet.cell(row=3, column=7).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=7).value = 20

            # Frequency
            present_sheet.cell(row=3, column=8).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=8).value = 20

            # Frequency
            present_sheet.cell(row=3, column=9).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=9).value = 20

            # Frequency
            present_sheet.cell(row=3, column=10).value = 25
            # Duty cycle
            present_sheet.cell(row=4, column=10).value = 20

            #------------------------------------------------------- IVVT & EVVT
            # Frequency
            present_sheet.cell(row=3, column=11).value = 250
            present_sheet.cell(row=4, column=11).value = 20

            # Frequency
            present_sheet.cell(row=3, column=17).value = 250
            # Duty cycle
            present_sheet.cell(row=4, column=17).value = 20

            #-------------------------------------------------------- LSHUP/DOWN
            # Frequency
            present_sheet.cell(row=3, column=12).value = 10
            present_sheet.cell(row=4, column=12).value = 25

            # Frequency
            present_sheet.cell(row=3, column=13).value = 10
            # Duty cycle
            present_sheet.cell(row=4, column=13).value = 25

            #----------------------------------------------------------- ALT_CMD
            # Frequency
            present_sheet.cell(row=3, column=14).value = 100
            # Duty cycle
            present_sheet.cell(row=4, column=14).value = 30

            #---------------------------------------------------------------- WG
            # Frequency
            present_sheet.cell(row=3, column=15).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=15).value = 30

            #---------------------------------------------------------------- CP
            # Frequency
            present_sheet.cell(row=3, column=16).value = 15
            # Duty cycle
            present_sheet.cell(row=4, column=16).value = 30

        #=======================================================================
        # Full Load Mode
        #=======================================================================

        if main_sheet.cell(row=2, column=2).value == 'full_load':
            #-------------------------------------------------------------- IGCx
            # Frequency
            present_sheet.cell(row=3, column=3).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=3).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=4).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=4).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=5).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=5).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=6).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=6).value = 100 - 87.58

            #-------------------------------------------------------------- INJx
            # Frequency
            present_sheet.cell(row=3, column=7).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=7).value = 25

            # Frequency
            present_sheet.cell(row=3, column=8).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=8).value = 25

            # Frequency
            present_sheet.cell(row=3, column=9).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=9).value = 25

            # Frequency
            present_sheet.cell(row=3, column=10).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=10).value = 25

            #------------------------------------------------------- IVVT & EVVT
            # Frequency
            present_sheet.cell(row=3, column=11).value = 250
            present_sheet.cell(row=4, column=11).value = 50

            # Frequency
            present_sheet.cell(row=3, column=17).value = 250
            # Duty cycle
            present_sheet.cell(row=4, column=17).value = 50

            #-------------------------------------------------------- LSHUP/DOWN
            # Frequency
            present_sheet.cell(row=3, column=12).value = 10
            present_sheet.cell(row=4, column=12).value = 50

            # Frequency
            present_sheet.cell(row=3, column=13).value = 10
            # Duty cycle
            present_sheet.cell(row=4, column=13).value = 50

            #----------------------------------------------------------- ALT_CMD
            # Frequency
            present_sheet.cell(row=3, column=14).value = 100
            # Duty cycle
            present_sheet.cell(row=4, column=14).value = 60

            #---------------------------------------------------------------- WG
            # Frequency
            present_sheet.cell(row=3, column=15).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=15).value = 70

            #---------------------------------------------------------------- CP
            # Frequency
            present_sheet.cell(row=3, column=16).value = 15
            # Duty cycle
            present_sheet.cell(row=4, column=16).value = 50

        #=======================================================================
        # Worst Case Mode
        #=======================================================================

        if main_sheet.cell(row=2, column=2).value == 'worst_case':
            #-------------------------------------------------------------- IGCx
            # Frequency
            present_sheet.cell(row=3, column=3).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=3).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=4).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=4).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=5).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=5).value = 100 - 87.58

            # Frequency
            present_sheet.cell(row=3, column=6).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=6).value = 100 - 87.58

            #-------------------------------------------------------------- INJx
            # Frequency
            present_sheet.cell(row=3, column=7).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=7).value = 25

            # Frequency
            present_sheet.cell(row=3, column=8).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=8).value = 25

            # Frequency
            present_sheet.cell(row=3, column=9).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=9).value = 25

            # Frequency
            present_sheet.cell(row=3, column=10).value = 54
            # Duty cycle
            present_sheet.cell(row=4, column=10).value = 25

            #------------------------------------------------------- IVVT & EVVT
            # Frequency
            present_sheet.cell(row=3, column=11).value = 50
            present_sheet.cell(row=4, column=11).value = 90

            # Frequency
            present_sheet.cell(row=3, column=17).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=17).value = 90

            #-------------------------------------------------------- LSHUP/DOWN
            # Frequency
            present_sheet.cell(row=3, column=12).value = 10
            present_sheet.cell(row=4, column=12).value = 70

            # Frequency
            present_sheet.cell(row=3, column=13).value = 10
            # Duty cycle
            present_sheet.cell(row=4, column=13).value = 70

            #----------------------------------------------------------- ALT_CMD
            # Frequency
            present_sheet.cell(row=3, column=14).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=14).value = 90

            #---------------------------------------------------------------- WG
            # Frequency
            present_sheet.cell(row=3, column=15).value = 50
            # Duty cycle
            present_sheet.cell(row=4, column=15).value = 90

            #---------------------------------------------------------------- CP
            # Frequency
            present_sheet.cell(row=3, column=16).value = 10
            # Duty cycle
            present_sheet.cell(row=4, column=16).value = 90

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=1).value = \
            main_sheet.cell(row=x + 8, column=1).value

            if main_sheet.cell(row=x + 8, column=2).value == 'LAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = BROWN_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

            if main_sheet.cell(row=x + 8, column=2).value == 'CAN Failure!':
                present_sheet.cell(row=x + 0, column=1).fill = VIOLET_FILL
                present_sheet.cell(row=1, column=1).fill = YELLOW_FILL

        present_sheet.cell(row=2, column=1).value = 'Alert'
        present_sheet.cell(row=3, column=1).value = 'Nominal Frequency [Hz]'
        present_sheet.cell(row=4, column=1).value = 'Nominal Duty [%]'

        # print("Service packs")
        for x in range(5, EXCEL_VERT_SCOPE - 7):
            present_sheet.cell(row=x + 0, column=3).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=181).value,
                               present_sheet, 3, x)

            present_sheet.cell(row=x + 0, column=4).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=182).value,
                               present_sheet, 4, x)

            present_sheet.cell(row=x + 0, column=5).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=183).value,
                               present_sheet, 5, x)

            present_sheet.cell(row=x + 0, column=6).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=184).value,
                               present_sheet, 6, x)

            present_sheet.cell(row=x + 0, column=7).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=185).value,
                               present_sheet, 7, x)

            present_sheet.cell(row=x + 0, column=8).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=186).value,
                               present_sheet, 8, x)

            present_sheet.cell(row=x + 0, column=9).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=187).value,
                               present_sheet, 9, x)

            present_sheet.cell(row=x + 0, column=10).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=188).value,
                               present_sheet, 10, x)

            present_sheet.cell(row=x + 0, column=11).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=189).value,
                               present_sheet, 11, x)

            present_sheet.cell(row=x + 0, column=12).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=190).value,
                               present_sheet, 12, x)

            present_sheet.cell(row=x + 0, column=13).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=191).value,
                               present_sheet, 13, x)

            present_sheet.cell(row=x + 0, column=14).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=174).value,
                               present_sheet, 14, x)

            present_sheet.cell(row=x + 0, column=15).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=193).value,
                               present_sheet, 15, x)

            present_sheet.cell(row=x + 0, column=16).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=175).value,
                               present_sheet, 16, x)

            present_sheet.cell(row=x + 0, column=17).value = \
            out_freq_service_pack(main_sheet.cell(row=x + 8, column=176).value,
                               present_sheet, 17, x)

#===============================================================================
# Frequency Output Fluctuations
#===============================================================================
        if __name__ == "__main__":
            t01 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(3,))
            t02 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(4,))
            t03 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(5,))
            t04 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(6,))
            t05 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(7,))
            t06 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(8,))
            t07 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(9,))
            t08 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(10,))
            t09 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(11,))
            t10 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(12,))
            t11 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(13,))
            t12 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(15,))
            t13 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(16,))
            t14 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(17,))
            t15 = threading.Thread(target=frequency_output_fluctuations,
                                  args=(14,))

            t01.start()
            t02.start()
            t03.start()
            t04.start()
            t05.start()
            t06.start()
            t07.start()
            t08.start()
            t09.start()
            t10.start()
            t11.start()
            t12.start()
            t13.start()
            t14.start()
            t15.start()

            t01.join()
            t02.join()
            t03.join()
            t04.join()
            t05.join()
            t06.join()
            t07.join()
            t08.join()
            t09.join()
            t10.join()
            t11.join()
            t12.join()
            t13.join()
            t14.join()
            t15.join()

#-------------------------------------------------------------------- Relocating

        #-------------------------------------------------------------- Coloring
        present_sheet.freeze_panes = 'B5'

        centr_align(present_sheet, 'A1:Q{}'.format(EXCEL_VERT_SCOPE))
        make_bold(present_sheet, 'A1:Q4')

        for column in Excel_std_column_names[:1]:
            present_sheet.column_dimensions[column].width = 24

        for column in Excel_std_column_names[1:2]:
            present_sheet.column_dimensions[column].width = 12

        for column in Excel_std_column_names[2:18]:
            present_sheet.column_dimensions[column].width = 18

        #=======================================================================
        # Finalizing
        #=======================================================================
        # print("Finalizing")
        #----------------------------------------------- EMI Frequency Insertion
        for sheet in workbook.sheetnames:
            if sheet != 'ETC Registers' \
            and sheet != 'Digital Inputs'\
            and sheet != 'Reset Registers'\
            and sheet != 'Monitoring Unit' \
            and sheet != 'TLE8888 Registers'\
            and sheet != 'Frequency Outputs' \
            and sheet != 'Temperature Values' \
            and sheet != workbook.sheetnames[0]:

                emi_frequencies_added_sp1(workbook[sheet])
                workbook[sheet].cell(row=1, column=9).value = 'EMI [MHz]'
                workbook[sheet].column_dimensions['I'].width = 8

        for x in range(1, EXCEL_VERT_SCOPE - 7):
            workbook['Digital Inputs'].cell(row=x + 1, column=2).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

            workbook['TLE8888 Registers'].cell(row=x, column=39).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

            workbook['ETC Registers'].cell(row=x, column=26).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

            workbook['Reset Registers'].cell(row=x, column=9).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

            workbook['Monitoring Unit'].cell(row=x, column=2).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

            workbook['Frequency Outputs'].cell(row=x, column=2).value = \
            workbook['KNK'].cell(row=x + 0, column=9).value

        workbook['Frequency Outputs'].cell(row=2, column=2).value = ''
        workbook['Frequency Outputs'].cell(row=3, column=2).value = ''
        workbook['Frequency Outputs'].cell(row=4, column=2).value = ''

        #------------------------------------------------- Formatting the fields
        # print("Formatting")
        main_sheet.freeze_panes = "C10"
        main_sheet.column_dimensions['A'].width = 14
        main_sheet.column_dimensions['B'].width = 21

        for column in Excel_std_column_names[2:128]:
            main_sheet.column_dimensions[column].width = 12

        for column in Excel_std_column_names[128:EXCEL_HRZN_SCOPE]:
            main_sheet.column_dimensions[column].width = 18

        centr_align(main_sheet, 'A8:GT{}'.format(EXCEL_VERT_SCOPE))
        make_blue(main_sheet, 'A1:B7')
        make_bold(main_sheet, 'A8:GT9')

        #-------------------------------------------------------- Correcting IGK
        for cell in workbook[workbook.sheetnames[1]].iter_rows(min_row=2,
                                             max_row=EXCEL_VERT_SCOPE,
                                             min_col=13,
                                             max_col=13):
            for j in range(len(cell)):
                if cell[j].value == 'LAN Failure!'\
                or cell[j].value == 'CAN Failure!':
                    cell[j].value = workbook[workbook.sheetnames[1]]\
                    .cell(row=2, column=2).value

        #---------------------------------------------------------------- Saving
        workbook.save(filename=output_file_name)
        workbook.close()
        os.remove(input_file_name)

        end_time = time.time()
        time_took_h = floor((end_time - start_time) / 3600)
        time_took_m = floor((end_time - start_time - 3600 * time_took_h) / 60)
        time_took_s = floor((end_time - start_time) % 60)
        time_now = str(datetime.datetime.now())[:-7]
        time_now = time_now[:10] + ' ' + time_now[10:]

        print_str = 'File Saved       =>   ' + output_file_name + '\n' + \
                    'Ended on         =>   [{}]\n'.format(time_now) + \
                    'Time Elapsed     =>   ' + \
                    '[{}h : {}m : {}s]'.format(time_took_h,
                                                  time_took_m,
                                                  time_took_s) + '\n\n'
        print(print_str)

        entries = list(os.scandir('TestReports'))
        if len(entries) == 0:
            break

    else:
        continue

very_end_time = time.time()
very_time_took_h = floor((very_end_time - very_start_time) / 3600)
very_time_took_m = floor((very_end_time - very_start_time - \
                          3600 * very_time_took_h) / 60)
very_time_took_s = floor((very_end_time - very_start_time) % 60)

#===============================================================================
# Holy Father Has Something to Say!
#===============================================================================
print('It is finished.')
last_print_str = 'Total Time Took  =>   [{}h : {}m : {}s]' \
.format(very_time_took_h, very_time_took_m, very_time_took_s)
print(last_print_str)
